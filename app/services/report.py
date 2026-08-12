"""MIS workbook generator — a formula-driven, board-ready Excel file.

Design: the raw revenue / expense / labour facts are written to data sheets;
every summary number on the P&L and dashboard sheets is an Excel ``SUMIFS`` (or
arithmetic) formula referencing those data sheets. Open the workbook, tweak a
data row, and every report number recalculates — nothing is hard-coded.
"""

from __future__ import annotations

import datetime as _dt
from dataclasses import dataclass
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from .. import config
from ..database import transaction
from .calc import (
    EXPENSE_TYPE_INDIRECT,
    EXPENSE_TYPE_PROFESSIONAL,
    MISData,
    MISOptions,
    compute,
    expense_type,
    financial_year,
    normalize_fy,
    revenue_category,
)
from .resolution import norm
from ..util import month_label, periods_label

# --- palette & formats -------------------------------------------------------

NAVY = "1F2A44"
BLUE = "2F7DF6"
LIGHT = "EAF0FB"
GREY = "F0F1F4"
# Indian lakh/crore grouping, right-sized per value magnitude so small numbers
# don't pick up phantom commas. Escaped commas (``\,``) in an Excel format
# string are literal — they render even when the preceding ``#`` placeholder
# has no digit to fill. The previous one-size-fits-crore format meant a
# ``1,500`` cell would show garbage commas on screen. Tiered conditional
# sections keep each magnitude clean:
#   - ≥ 1 crore (10000000) → ``1,23,45,678``
#   - ≥ 1 lakh (100000)    → ``1,23,456``
#   - < 1 lakh             → ``1,500`` (standard 3-digit grouping)
# Excel allows max 3 conditional numeric sections, so big-negative red
# styling was dropped — negatives render in regular grouping with a minus.
INR = (r'[>=10000000]##\,##\,##\,##0;'
       r'[>=100000]##\,##\,##0;'
       r'##,##0')
PCT = '0.0%'
HOURS = '#,##0.0'

_TITLE = Font(name="Segoe UI", size=15, bold=True, color=NAVY)
_SUB = Font(name="Segoe UI", size=10, italic=True, color="5B6577")
_HEAD = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
_BOLD = Font(name="Segoe UI", size=10, bold=True)
_NORMAL = Font(name="Segoe UI", size=10)
_KPI = Font(name="Segoe UI", size=18, bold=True, color=NAVY)

_HEAD_FILL = PatternFill("solid", fgColor=NAVY)
_SUBHEAD_FILL = PatternFill("solid", fgColor="3D4E70")
_TOTAL_FILL = PatternFill("solid", fgColor=LIGHT)
_SECTION_FILL = PatternFill("solid", fgColor="DEE5F1")
_KPI_FILL = PatternFill("solid", fgColor=GREY)


# Heuristic — what counts as "Reimbursement / OPE" income vs Sales income.
# Used to populate the Category column on the Revenue data sheet so the
# Partner-Manager P&L can split "Sales" from "Reimb & OPE" via SUMIFS.


def _service_category(name: str) -> str:
    # Canonical classifier lives in calc.revenue_category so the Excel
    # Category column and the calc roll-ups never diverge.
    return revenue_category(name)
_thin = Side(style="thin", color="D2D6DE")
_BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
_CENTER = Alignment(horizontal="center", vertical="center")
_RIGHT = Alignment(horizontal="right")


def _q(sheet: str) -> str:
    """Quote a sheet name for use inside a formula."""
    return f"'{sheet}'"


# --- generic cell helpers ----------------------------------------------------

def _cell(ws, row, col, value, *, font=_NORMAL, fill=None, fmt=None,
          align=None, border=False):
    c = ws.cell(row=row, column=col, value=value)
    c.font = font
    if fill:
        c.fill = fill
    if fmt:
        c.number_format = fmt
    if align:
        c.alignment = align
    if border:
        c.border = _BORDER
    return c


def _header_row(ws, row, headers, start_col=1):
    for i, text in enumerate(headers):
        _cell(ws, row, start_col + i, text, font=_HEAD, fill=_HEAD_FILL,
              align=_CENTER, border=True)


# --- master label maps -------------------------------------------------------

def _labels() -> dict:
    with transaction() as conn:
        cc = {r["id"]: r["code"]
              for r in conn.execute("SELECT id, code FROM cost_centres")}
        cc_active = [dict(r) for r in conn.execute(
            "SELECT id, code, name, cc_type FROM cost_centres "
            "WHERE active = 1 ORDER BY cc_type, code")]
        mgr = {r["id"]: r["code"]
               for r in conn.execute("SELECT id, code FROM managers")}
        ent = {r["id"]: r["name"]
               for r in conn.execute("SELECT id, name FROM entities")}
        svc = {r["id"]: r["name"]
               for r in conn.execute("SELECT id, name FROM services")}
        cli = {r["id"]: r["canonical_name"]
               for r in conn.execute("SELECT id, canonical_name FROM clients")}
    return {"cc": cc, "cc_active": cc_active, "mgr": mgr, "ent": ent,
            "svc": svc, "cli": cli}


# =============================== generation =================================

def generate(data: MISData, path: str | Path,
             windows: "Windows | None" = None) -> Path:
    """Write the full MIS workbook for *data* to *path*.

    *windows* are the three context periods shown alongside the selected
    month — see :func:`build_windows`, which builds them when the caller
    doesn't (passing them in avoids recomputing them for the HTML
    dashboard).
    """
    lbl = _labels()
    wb = Workbook()
    wb.remove(wb.active)

    if windows is None:
        windows = build_windows(data.options)
    fy_prior, ytd, last_year = (windows.fy_prior, windows.ytd,
                                windows.last_year)
    # The FY-prior window's revenue / expense rows are merged INTO the
    # "Revenue" / "Expenses" sheets under Scope = "FY Prior" (v0.3.118);
    # salary / reimbursements / provisions keep their " (FY)" sheets. The
    # two year-on-year windows get a full sheet set each — they OVERLAP
    # the other two (the year to date contains the selected months), so
    # merging them would double-count.
    fy_data = fy_prior.data

    _sheet_cover(wb, data, windows)
    _sheet_dashboard(wb, data)
    _sheet_budget_monthly(wb, data, lbl, fy_prior)
    rows_pl = _sheet_cost_centre(wb, data, lbl)
    _sheet_partner_manager(wb, data, lbl, prior=fy_prior)
    _sheet_pm_comparison(wb, data, lbl, windows)
    _sheet_entity(wb, data, lbl)
    _sheet_service(wb, data, lbl)
    _sheet_client_billing(wb, data, lbl, fy_prior)
    _sheet_employee_register(wb, data, lbl)
    _sheet_client_register(wb, data, lbl, fy_prior)
    _sheet_comparatives(wb, data, lbl, windows, rows_pl)
    # Revenue / Expenses carry BOTH windows (v0.3.118): the FY-prior rows
    # first (they are the earlier months), then the selected period's,
    # told apart by the Scope column.
    _sheet_revenue(wb, data, lbl, fy_data=fy_data)
    _sheet_expenses(wb, data, lbl, fy_data=fy_data)
    _sheet_salary(wb, data, lbl)
    _sheet_reimbursements(wb, data, lbl)
    _sheet_provisions(wb, data, lbl)
    if fy_data is not None:
        # Revenue / Expenses for this window are already on the main
        # sheets (Scope = "FY Prior"); these three keep their own.
        _sheet_salary(wb, fy_data, lbl, FYS)
        _sheet_reimbursements(wb, fy_data, lbl, FYS)
        _sheet_provisions(wb, fy_data, lbl, FYS)
    # Year-on-year: the financial year to date, and the same span a year
    # earlier — one full sheet set each, driving the "(Cmp)" P&L and the
    # Comparatives sheet.
    for win in (ytd, last_year):
        if win.data is None:
            continue
        _sheet_revenue(wb, win.data, lbl, win.suffix, scope=win.label)
        _sheet_expenses(wb, win.data, lbl, win.suffix, scope=win.label)
        _sheet_salary(wb, win.data, lbl, win.suffix)
        _sheet_reimbursements(wb, win.data, lbl, win.suffix)
        _sheet_provisions(wb, win.data, lbl, win.suffix)

    # Dashboard KPIs + Cover totals link to the Cost Centre P&L total row.
    _link_dashboard(wb, rows_pl)
    _link_cover(wb, rows_pl)

    path = Path(path)
    wb.save(path)
    return path


# Data-sheet suffixes, one per context window (v0.3.121).
FYS = " (FY)"     # the FY months BEFORE the selected period
YTD = " (YTD)"    # the financial year to date, through the selected month
LY = " (LY)"      # the same year-to-date span one financial year earlier

# Companion data sheet behind the "Budget vs Monthly Sales" summary
# (v0.3.115): one row per (partner, FY-to-date month, sales) so that
# sheet's month columns can be live SUMIFS instead of baked values.
BUDGET_DATA_SHEET = "Budget Sales (Monthly)"

# --- merged Revenue / Expenses data sheets (v0.3.118) ------------------------
#
# "Revenue (FY)" / "Expenses (FY)" used to be separate sheets holding the
# FY-prior window (the months of the financial year before the selected
# period — see :func:`_fy_prior_periods`). They are now written into the
# SAME "Revenue" / "Expenses" sheets, so the operator reads one continuous
# transaction register per kind instead of hopping between two. A trailing
# **Scope** column says which window each row belongs to.
#
# Two invariants make this safe:
#
# 1. The two windows never share a month. ``_fy_prior_periods`` returns
#    months strictly BEFORE the earliest selected month (or the whole
#    previous FY when the selection starts in April), so a Period label
#    identifies its window on its own. Formulas already keyed on Period
#    (Client Billing, the Budget data sheet, the Employee Register's
#    office-indirect pool) therefore need NO scope criterion — and got
#    simpler, since both windows now live on one sheet.
# 2. Partner / cost-centre aggregates DO need one, and the current-window
#    criterion is ``<>FY Prior`` rather than ``=Current``: a row the
#    operator types in below the generated data has an empty Scope cell,
#    and SUMIFS ``<>`` criteria match empty cells — so an appended row
#    still counts towards the selected period, preserving the v0.3.117
#    "everything I added counts" guarantee. Tagging it "FY Prior" by hand
#    is the (discoverable) way to book it to the earlier window instead.
#
# The year-on-year windows ( (YTD) / (LY) ) stay genuinely separate sheets:
# the year to date CONTAINS the selected months, so merging it in would
# double-count every row it shares with them.
SCOPE_HEADER = "Scope"
SCOPE_CURRENT = "Current"
SCOPE_FY = "FY Prior"
SCOPE_COL_REV = "K"   # Revenue  sheet: A..J data + K Scope
SCOPE_COL_EXP = "M"   # Expenses sheet: A..L data + M Scope

# Suffixes whose window has its own Revenue / Expenses sheet rather than
# sharing the main one under a Scope value.
_OWN_SHEET_SUFFIXES = (YTD, LY)


def _data_sheet_q(base: str, sfx: str = "") -> str:
    """Quoted name of the Revenue / Expenses data sheet to read for *sfx*.

    ``FYS`` shares the main sheet (the FY-prior window is a Scope column,
    not a sheet of its own); the year-on-year windows have their own.
    """
    return _q(base + (sfx if sfx in _OWN_SHEET_SUFFIXES else ""))


def _scope_crit(sheet_q: str, scope_col: str, sfx: str = "") -> str:
    """The extra SUMIFS criteria pair restricting a merged data sheet to
    one period window — ``""`` (selected period) or :data:`FYS`.

    Returns a leading-comma fragment to splice into a SUMIFS call, or an
    empty string for a window with its own sheet (nothing to restrict).
    """
    if sfx in _OWN_SHEET_SUFFIXES:
        return ""
    value = SCOPE_FY if sfx == FYS else f"<>{SCOPE_FY}"
    return f',{sheet_q}!${scope_col}:${scope_col},"{value}"'


# --- the context windows (v0.3.121) ------------------------------------------

@dataclass
class Window:
    """One span of months shown alongside the selected reporting period.

    ``suffix`` names the data sheets its figures are read from, so one set
    of formula builders serves every window.
    """
    periods: list[str]
    suffix: str
    data: MISData | None

    @property
    def active(self) -> bool:
        return bool(self.periods) and self.data is not None

    @property
    def label(self) -> str:
        """Human label for the whole window ("Apr-26 to Jun-26")."""
        return periods_label(self.periods) if self.periods else "—"


@dataclass
class Windows:
    """The three context spans every MIS run carries (v0.3.121).

    The operator picks the reporting month(s) and nothing else — the
    "Compare with" selection is gone, and these are derived:

    * **fy_prior** — the financial-year months BEFORE the selected period
      (Jun-26 → Apr-26..May-26). One cumulative column at the end of each
      partner block on the Partner-Manager P&L, one column per month on
      Client Billing, and one row per month on the Client Register. Read
      from the ``Scope = "FY Prior"`` rows of Revenue / Expenses plus the
      " (FY)" Salary / Reimbursements / Provisions sheets.
    * **ytd** — the financial year to date, INCLUDING the selected months
      (Jun-26 → Apr-26..Jun-26).
    * **last_year** — the same span one financial year earlier (Apr-25..
      Jun-25).

    The last two are the year-on-year pair behind the "Partner-Manager
    P&L (Cmp)" and "Comparatives" sheets, and each has its own full set
    of data sheets: the year to date CONTAINS the selected months, so
    merging it into the main sheets would double-count.
    """
    selected: list[str]
    fy_prior: Window
    ytd: Window
    last_year: Window


def build_windows(options: MISOptions) -> Windows:
    """Derive the three context windows for a run (v0.3.121).

    Each is computed with the same calc engine as the reporting period,
    so every figure in the workbook stays a live SUMIFS.
    """
    def run(periods: list[str], suffix: str) -> Window:
        data = compute(MISOptions(
            periods=periods,
            include_reimbursement=options.include_reimbursement,
            location_ids=options.location_ids)) if periods else None
        return Window(periods, suffix, data)

    ytd = _fy_to_date_periods(options.periods)
    return Windows(
        selected=sorted(options.periods),
        fy_prior=run(_fy_prior_periods(options.periods), FYS),
        ytd=run(ytd, YTD),
        last_year=run(_shift_years(ytd, -1), LY))


def _fy_to_date_periods(periods: list[str]) -> list[str]:
    """The financial year to date THROUGH the latest selected month.

    Jun-26 → Apr-26..Jun-26; Nov-26..Feb-27 → Apr-26..Feb-27. Unlike
    :func:`_fy_prior_periods` this INCLUDES the selected months, and it
    fills any gap a non-contiguous selection leaves (Apr-26 + Jun-26 →
    Apr-26..Jun-26), so the year-to-date column is a true year to date.
    """
    if not periods:
        return []
    latest = max(periods)
    return _fy_months_through(financial_year(latest), latest)


def _shift_years(periods: list[str], years: int) -> list[str]:
    """Shift every month by whole years — the same span, another FY."""
    out = []
    for p in periods:
        try:
            out.append(f"{int(p[:4]) + years:04d}-{p[5:7]}")
        except (ValueError, IndexError):
            continue
    return out


def _fy_prior_periods(periods: list[str]) -> list[str]:
    """The FY months BEFORE the earliest selected month (v0.3.102).

    MIS for Jul-26 → Apr-26..Jun-26; MIS for Nov-26..Feb-27 → Apr-26..
    Oct-26 (anchored on the EARLIEST selected month's financial year).
    Edge case per the operator: a selection starting in April has no
    prior FY months — show the WHOLE previous FY instead (Apr..Mar).
    Empty input → empty output.
    """
    if not periods:
        return []
    earliest = min(periods)
    try:
        start_year = int(financial_year(earliest).split("-")[0])
    except (ValueError, IndexError):
        return []
    fy_start = f"{start_year:04d}-04"
    if earliest == fy_start:
        # First month of the FY — cumulative = the whole previous FY.
        return ([f"{start_year - 1:04d}-{m:02d}" for m in range(4, 13)]
                + [f"{start_year:04d}-{m:02d}" for m in range(1, 4)])
    months = []
    y, m = start_year, 4
    while f"{y:04d}-{m:02d}" < earliest:
        months.append(f"{y:04d}-{m:02d}")
        y, m = (y + 1, 1) if m == 12 else (y, m + 1)
    return months


# --- Cover -------------------------------------------------------------------

# First row of the Cover's label/value block. _link_cover scans from here
# for the three headline totals, so both must agree.
COVER_FIRST_ROW = 6


def _sheet_cover(wb: Workbook, data: MISData,
                 windows: "Windows | None" = None) -> None:
    ws = wb.create_sheet("Cover")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 70

    _cell(ws, 2, 2, config.ORG_NAME, font=Font(size=20, bold=True, color=NAVY))
    _cell(ws, 3, 2, "Management Information System",
          font=Font(size=13, color=BLUE))
    periods = periods_label(data.options.periods)
    rows = [
        ("Reporting period(s)", periods),
        # The three derived context windows (v0.3.121).
        ("Previous months this FY",
         windows.fy_prior.label if windows else "—"),
        ("Year to date", windows.ytd.label if windows else "—"),
        ("Same period last year",
         windows.last_year.label if windows else "—"),
        ("Generated on", _dt.date.today().strftime("%d %b %Y")),
        ("Reimbursements in MIS",
         "Included" if data.options.include_reimbursement else "Excluded"),
        ("Locations considered", data.location_label),
        ("Office overhead basis",
         "Office indirect expenses ÷ (partner-team employees + partners)"),
        # The three totals become live formulas via _link_cover once the
        # Cost Centre P&L sheet exists (v0.3.99); 0 is just a placeholder.
        ("Total revenue", 0),
        ("Total cost", 0),
        ("Net profit", 0),
    ]
    for i, (k, v) in enumerate(rows):
        r = COVER_FIRST_ROW + i
        c = ws.cell(row=r, column=2, value=k)
        c.font = _BOLD
        c.alignment = Alignment(horizontal="left")
        d = ws.cell(row=r, column=3, value=v)
        d.font = _NORMAL
    ws.column_dimensions["C"].width = 32
    if data.warnings:
        notes_row = COVER_FIRST_ROW + len(rows) + 1
        _cell(ws, notes_row, 2, "Notes:", font=_BOLD)
        for i, w in enumerate(data.warnings):
            _cell(ws, notes_row + 1 + i, 2, "• " + w, font=_SUB)


# --- Dashboard ---------------------------------------------------------------

def _sheet_dashboard(wb: Workbook, data: MISData) -> None:
    ws = wb.create_sheet("Dashboard")
    ws.sheet_view.showGridLines = False
    for col, width in (("A", 3), ("B", 26), ("C", 22), ("D", 26), ("E", 22)):
        ws.column_dimensions[col].width = width

    _cell(ws, 2, 2, "MIS Dashboard", font=_TITLE)
    _cell(ws, 3, 2, "Period(s): " + periods_label(data.options.periods), font=_SUB)

    # KPI tiles — values filled later by _link_dashboard.
    tiles = [("Total Revenue", 5, 2), ("Total Cost", 5, 4),
             ("Net Profit", 8, 2), ("Net Margin", 8, 4)]
    for label, r, c in tiles:
        _cell(ws, r, c, label, font=_BOLD, fill=_KPI_FILL)
        ws.cell(row=r, column=c + 1).fill = _KPI_FILL
        kpi = _cell(ws, r + 1, c, 0, font=_KPI, fill=_KPI_FILL, fmt=INR)
        ws.cell(row=r + 1, column=c + 1).fill = _KPI_FILL
        ws.merge_cells(start_row=r + 1, start_column=c,
                       end_row=r + 1, end_column=c + 1)
        kpi.alignment = Alignment(horizontal="left", vertical="center")

    _cell(ws, 12, 2, "See 'Cost Centre P&L' for the full partner-wise "
          "profitability. Total Revenue includes Reimbursements (OPE).",
          font=_SUB)


def _link_cover(wb: Workbook, rows_pl: dict) -> None:
    """Point the Cover's three headline totals at the Cost Centre P&L
    total row (v0.3.99 — they were baked fmt_inr strings before), so the
    Cover stays true when the operator edits a data row."""
    ws = wb["Cover"]
    pl = _q("Cost Centre P&L")
    total = rows_pl["total_row"]
    ti = rows_pl["col_total_income"]
    prof = rows_pl["col_profit"]
    formulas = {
        "Total revenue": f"={pl}!{ti}{total}",
        "Total cost": f"={pl}!{ti}{total}-{pl}!{prof}{total}",
        "Net profit": f"={pl}!{prof}{total}",
    }
    for r in range(COVER_FIRST_ROW, COVER_FIRST_ROW + 30):
        label = ws.cell(row=r, column=2).value
        if label in formulas:
            cell = ws.cell(row=r, column=3, value=formulas.pop(label))
            cell.font = _NORMAL
            cell.number_format = INR
        if not formulas:
            break


def _link_dashboard(wb: Workbook, rows_pl: dict) -> None:
    ws = wb["Dashboard"]
    pl = _q("Cost Centre P&L")
    total = rows_pl["total_row"]
    ti = rows_pl["col_total_income"]
    prof = rows_pl["col_profit"]
    # (row, col): (formula, number_format). Total Revenue = Total Income;
    # Total Cost = Total Income − Net Profit (no single Total-Cost column now).
    mapping = {
        (6, 2): (f"={pl}!{ti}{total}", INR),
        (6, 4): (f"={pl}!{ti}{total}-{pl}!{prof}{total}", INR),
        (9, 2): (f"={pl}!{prof}{total}", INR),
        (9, 4): (f"={pl}!{rows_pl['col_margin']}{total}", PCT),
    }
    for (r, c), (formula, fmt) in mapping.items():
        cell = ws.cell(row=r, column=c, value=formula)
        cell.font = _KPI
        cell.number_format = fmt


# --- Budget vs Monthly Sales -------------------------------------------------

def _fy_months_through(fy: str, end_period: str) -> list[str]:
    """All FY months (Apr-start to Mar-end) up to and including *end_period*.

    Indian financial year, e.g. ``fy='2025-26'`` → ``['2025-04', ..., '2026-03']``.
    Truncated to *end_period* (so a Jan run shows Apr-Jan, not the empty
    Feb-Mar months ahead).
    """
    try:
        start_year = int(fy.split('-')[0])
    except (ValueError, IndexError):
        return []
    months = [f"{start_year:04d}-{m:02d}" for m in range(4, 13)]
    months += [f"{start_year + 1:04d}-{m:02d}" for m in range(1, 4)]
    if end_period in months:
        return months[:months.index(end_period) + 1]
    return months


def _month_short(period: str) -> str:
    """``'2025-04'`` → ``'Apr-25'`` — delegates to :func:`util.month_label`
    so every sheet renders periods identically (cross-sheet SUMIFS match
    on these labels)."""
    return month_label(period)


def budget_monthly_data(data: MISData, lbl: dict) -> dict | None:
    """FY-to-date monthly partner sales + annual budgets — the single
    source of truth behind both the Excel 'Budget vs Monthly Sales' sheet
    and the HTML dashboard's budget section, so the two never disagree.

    Returns ``None`` when there's nothing to show (no period / no partner).
    Otherwise a dict: ``fy``, ``months`` (FY-to-date), ``partners`` (the
    active partner rows), ``monthly`` (code → {period: sales}), ``budgets``
    (code → annual target), ``remaining_months`` (months left in the FY).
    """
    if not data.options.periods:
        return None
    latest = max(data.options.periods)
    fy = financial_year(latest)
    months = _fy_months_through(fy, latest)
    partners = [c for c in lbl["cc_active"] if c["cc_type"] == "partner"]
    if not months or not partners:
        return None

    placeholders = ','.join('?' * len(months))
    svc_labels = lbl["svc"]
    with transaction() as conn:
        # Per service so we can keep SALES only — reimbursement / OPE
        # recoveries also come through the Sales Register but are NOT sales
        # (v0.3.89): exclude them from the budget-vs-sales view.
        sales_rows = conn.execute(
            f"SELECT cc.code AS code, v.period AS period, "
            f"       s.service_id AS service_id, "
            f"       COALESCE(SUM(s.amount), 0) AS amount "
            f"FROM voucher_splits s "
            f"JOIN vouchers v ON v.id = s.voucher_id "
            f"JOIN cost_centres cc ON cc.id = s.cost_centre_id "
            f"WHERE v.kind = 'sales' AND v.period IN ({placeholders}) "
            f"  AND cc.cc_type = 'partner' "
            f"GROUP BY cc.code, v.period, s.service_id",
            months).fetchall()
        # Match on the NORMALISED financial year so loosely-typed master
        # values ('2026 - 27', '2026/27', …) still resolve to this FY —
        # an exact match left the Annual Budget column reading 0.
        budget_rows = conn.execute(
            "SELECT cc.code AS code, t.financial_year AS fy, "
            "       t.target_amount AS amount "
            "FROM targets t "
            "JOIN cost_centres cc ON cc.id = t.cost_centre_id").fetchall()

    monthly: dict[str, dict[str, float]] = {}
    for row in sales_rows:
        if revenue_category(svc_labels.get(row["service_id"])) != "Income":
            continue   # skip reimbursement / OPE / round-off ledgers
        m = monthly.setdefault(row["code"], {})
        m[row["period"]] = m.get(row["period"], 0.0) + row["amount"]
    fy_norm = normalize_fy(fy)
    budgets = {row["code"]: row["amount"] for row in budget_rows
               if normalize_fy(row["fy"]) == fy_norm}
    return {
        "fy": fy, "months": months, "partners": partners,
        "monthly": monthly, "budgets": budgets,
        "remaining_months": max(0, 12 - len(months)),
    }


def _sheet_budget_monthly(wb: Workbook, data: MISData, lbl: dict,
                          prior: "Window | None" = None) -> None:
    """Year-to-date monthly sales per partner cost centre, vs annual budget.

    Independent of the selected MIS period: always shows the full FY-to-date
    picture so the board sees trend context alongside the headline P&L.

    Every amount cell is a live formula, all the way down (v0.3.115/116):
    each month cell SUMIFS the companion data sheet
    (:data:`BUDGET_DATA_SHEET`) by partner code + month label, and that
    sheet's Sales cells are THEMSELVES SUMIFS over a transaction-level
    revenue register. The "Revenue" sheet covers the selected months plus
    (when the window is the FY to date) the earlier FY ones since
    v0.3.118; an operator-chosen comparison window lives on
    "Revenue (Cmp)" instead, and those months read that sheet (v0.3.120)
    — so a sales row edited or added post-generation flows through the
    data sheet into the month cell, its YTD, Variance and Average. A
    month covered by no window falls back to a baked value. YTD /
    Variance / Average were always formulas.
    """
    bm = budget_monthly_data(data, lbl)
    if bm is None:
        return
    fy = bm["fy"]
    months = bm["months"]
    partners = bm["partners"]
    monthly = bm["monthly"]
    budgets = bm["budgets"]

    ws = wb.create_sheet("Budget vs Monthly Sales")
    ws.sheet_view.showGridLines = False

    ws.column_dimensions["A"].width = 10  # Code
    ws.column_dimensions["B"].width = 26  # Cost Centre
    ws.column_dimensions["C"].width = 18  # Annual Budget
    for i, _ in enumerate(months):
        ws.column_dimensions[get_column_letter(4 + i)].width = 13
    ytd_col = 4 + len(months)
    var_col = ytd_col + 1
    avg_col = var_col + 1
    for col in (ytd_col, var_col, avg_col):
        ws.column_dimensions[get_column_letter(col)].width = 16

    _cell(ws, 1, 1, f"Budget vs Monthly Sales  —  FY {fy}", font=_TITLE)
    _cell(ws, 2, 1, "Year-to-date sales by partner cost centre. Annual budget "
                    "is read from the Targets master. The last column is the "
                    "average monthly sales still needed to hit budget "
                    "(Variance vs Budget ÷ months left in the FY).",
          font=_SUB)

    # Months still to run in this FY (Apr–Mar = 12). The last column shows
    # the average monthly sales each partner must still book to close the
    # gap to budget: Variance vs Budget ÷ remaining months.
    remaining_months = max(0, 12 - len(months))
    avg_header = (f"Avg / Remaining Month ({remaining_months})"
                  if remaining_months else "Avg / Remaining Month")
    headers = (["Code", "Cost Centre", "Annual Budget"]
               + [_month_short(m) for m in months]
               + ["YTD Total", "Variance vs Budget", avg_header])
    hrow = 4
    _header_row(ws, hrow, headers)

    body_start = hrow + 1
    r = body_start
    first_m = get_column_letter(4)
    last_m = get_column_letter(4 + len(months) - 1)
    ytd_L = get_column_letter(ytd_col)
    var_L = get_column_letter(var_col)
    bd = _q(BUDGET_DATA_SHEET)

    def avg_formula(row: int) -> str | float:
        # Variance vs Budget spread over the months left in the FY — the
        # run-rate needed to still hit the annual budget. No months left
        # (full FY generated) → nothing to spread, so 0.
        if remaining_months <= 0:
            return 0
        return f"={var_L}{row}/{remaining_months}"

    for partner in partners:
        code = partner["code"]
        _cell(ws, r, 1, code, border=True)
        _cell(ws, r, 2, partner["name"], border=True)
        _cell(ws, r, 3, round(budgets.get(code, 0.0), 2),
              fmt=INR, border=True)
        # Each month is a live SUMIFS over the companion data sheet, keyed
        # on the partner CODE (column A of this row) and the month LABEL
        # (this column's header cell) — so editing a sales row on that
        # sheet recomputes the month cell, the YTD, variance and average.
        for i, _period in enumerate(months):
            col = 4 + i
            col_L = get_column_letter(col)
            _cell(ws, r, col,
                  f'=SUMIFS({bd}!$C:$C,{bd}!$A:$A,$A{r},'
                  f'{bd}!$B:$B,{col_L}${hrow})',
                  fmt=INR, border=True)
        _cell(ws, r, ytd_col, f"=SUM({first_m}{r}:{last_m}{r})",
              font=_BOLD, fill=_TOTAL_FILL, fmt=INR, border=True)
        _cell(ws, r, var_col, f"=C{r}-{ytd_L}{r}", fmt=INR, border=True)
        _cell(ws, r, avg_col, avg_formula(r), fmt=INR, border=True)
        r += 1

    last_body = r - 1
    _cell(ws, r, 1, "", fill=_TOTAL_FILL, border=True)
    _cell(ws, r, 2, "TOTAL", font=_BOLD, fill=_TOTAL_FILL, border=True)
    for col in range(3, avg_col + 1):
        L = get_column_letter(col)
        if col == avg_col:
            formula = avg_formula(r)
        else:
            formula = f"=SUM({L}{body_start}:{L}{last_body})"
        _cell(ws, r, col, formula, font=_BOLD, fill=_TOTAL_FILL,
              fmt=INR, border=True)

    _subtotal_filter(ws, hrow, body_start, last_body,
                     list(range(3, avg_col + 1)), last_col=avg_col)
    ws.freeze_panes = f"D{body_start}"

    # Companion data sheet the month columns SUMIFS over. One row per
    # (partner, FY-to-date month), and each Sales cell is ITSELF a live
    # SUMIFS over a transaction-level revenue register (v0.3.116) —
    # partner code (col D) + Category "Income" (col I, so reimbursement /
    # OPE recoveries stay excluded like always) + Period label (col J).
    # Since v0.3.118 the "Revenue" sheet holds both the selected months
    # and the earlier FY ones (what used to be "Revenue (FY)"), and a
    # month belongs to exactly one window — so the Period label alone
    # selects the right rows and no Scope criterion is needed. A month
    # that only exists on the comparison register reads "Revenue (Cmp)"
    # (v0.3.120). Zero months get a row too, so a sales row ADDED
    # post-generation is still captured. A month covered by no register
    # keeps a baked value, and only when non-zero.
    source = {p: _q("Revenue") for p in data.options.periods}
    if prior is not None and prior.periods:
        prior_sheet = _data_sheet_q("Revenue", prior.suffix)
        for p in prior.periods:
            source.setdefault(p, prior_sheet)
    data_rows = []
    for partner in partners:
        code = partner["code"]
        pmonthly = monthly.get(code, {})
        for period in months:
            label = _month_short(period)
            rev = source.get(period)
            if rev:
                data_rows.append([
                    code, label,
                    f'=SUMIFS({rev}!$H:$H,{rev}!$D:$D,"{code}",'
                    f'{rev}!$I:$I,"Income",{rev}!$J:$J,"{label}")'])
            else:
                amt = pmonthly.get(period, 0.0)
                if abs(amt) >= 0.005:
                    data_rows.append([code, label, round(amt, 2)])
    _write_data_sheet(
        wb, BUDGET_DATA_SHEET,
        ["Partner", "Period", "Sales"],
        [12, 12, 16], data_rows)


# --- Cost Centre P&L (the core sheet) ----------------------------------------

def _sheet_cost_centre(wb: Workbook, data: MISData, lbl: dict) -> dict:
    ws = wb.create_sheet("Cost Centre P&L")
    ws.sheet_view.showGridLines = False
    # v0.3.96: columns now mirror the Partner-Manager P&L line structure —
    # Salary split billable / non-billable, Direct Expense split into
    # Professional Fees / Indirect Expenses / Provisions, plus Total Direct
    # Costs, Gross Profit and Gross Profit %. Gross/Net % are on sales
    # income (Revenue), matching the PM P&L.
    #   C Revenue  D Reimb(OPE)  E Total Income  F Salary(bill)
    #   G Salary(non-bill)  H Professional Fees  I Reimbursement Expenses
    #   J Provisions  K Total Direct Costs  L Gross Profit  M Gross %
    #   N Office Overhead  O Indirect Expenses  P Net Profit  Q Net %
    headers = ["Code", "Cost Centre", "Revenue", "Reimbursements (OPE)",
               "Total Income", "Salary (billable)", "Salary (non-billable)",
               "Professional Fees", "Reimbursement Expenses", "Provisions",
               "Total Direct Costs", "Gross Profit", "Gross Profit %",
               "Office Overhead", "Indirect Expenses", "Net Profit",
               "Net Profit %"]
    widths = [10, 26, 15, 17, 15, 15, 17, 15, 17, 13, 16, 14, 12, 15, 15, 14,
              12]
    for i, w in enumerate(widths):
        ws.column_dimensions[get_column_letter(1 + i)].width = w

    _cell(ws, 1, 1, "Cost Centre Profitability", font=_TITLE)
    _cell(ws, 2, 1, "Period(s): " + periods_label(data.options.periods), font=_SUB)
    hrow = 4
    _header_row(ws, hrow, headers)

    # All active cost centres (partners first, Office last), plus an
    # 'Unassigned' line if any fact lacks a cost centre.
    partners = [c for c in lbl["cc_active"] if c["cc_type"] != "office"]
    office = [c for c in lbl["cc_active"] if c["cc_type"] == "office"]
    has_unassigned = any(c.cost_centre_id is None for c in data.cost_centres)

    rev = _q("Revenue")
    exp = _q("Expenses")
    lab = _q("Salary")
    reimb = _q("Reimbursements")
    prov = _q("Provisions")
    # Revenue / Expenses also carry the FY-prior rows (v0.3.118) — this
    # sheet reports the SELECTED period, so every SUMIFS over them adds
    # the current-window criterion.
    rev_s = _scope_crit(rev, SCOPE_COL_REV)
    exp_s = _scope_crit(exp, SCOPE_COL_EXP)
    first = hrow + 1
    n_partners = len(partners)
    # Row order: partners, then Office, then (optional) Unassigned.
    n = n_partners + len(office) + (1 if has_unassigned else 0)
    last = first + n - 1
    total_row = last + 1

    def write_row(r, code, name, kind):
        _cell(ws, r, 1, code, font=_NORMAL, border=True)
        _cell(ws, r, 2, name, font=_NORMAL, border=True)
        # Reimbursements (OPE) income: Revenue rows whose Category (col I)
        # is Reimbursement or OPE (Amount=H, CostCentre=D).
        _cell(ws, r, 4,
              f'=SUMIFS({rev}!$H:$H,{rev}!$D:$D,$A{r},{rev}!$I:$I,'
              f'"Reimbursement"{rev_s})+SUMIFS({rev}!$H:$H,{rev}!$D:$D,$A{r},'
              f'{rev}!$I:$I,"OPE"{rev_s})',
              fmt=INR, border=True)
        # Revenue (sales income only) = all revenue for the CC minus OPE.
        _cell(ws, r, 3,
              f"=SUMIFS({rev}!$H:$H,{rev}!$D:$D,$A{r}{rev_s})-D{r}",
              fmt=INR, border=True)
        # Total Income = Revenue + Reimbursements (OPE).
        _cell(ws, r, 5, f"=C{r}+D{r}", fmt=INR, border=True)
        # Salary (billable) / (non-billable): Salary-type rows split on the
        # Billable column (K). Salary Amount=I, Charged To=C, Type=J.
        _cell(ws, r, 6,
              f'=SUMIFS({lab}!$I:$I,{lab}!$C:$C,$A{r},'
              f'{lab}!$J:$J,"Salary",{lab}!$K:$K,"Yes")',
              fmt=INR, border=True)
        _cell(ws, r, 7,
              f'=SUMIFS({lab}!$I:$I,{lab}!$C:$C,$A{r},'
              f'{lab}!$J:$J,"Salary",{lab}!$K:$K,"No")',
              fmt=INR, border=True)
        # Professional Fees: Expenses with Type = Professional Fees
        # (Expenses Amount=J, CostCentre=E, Type of Expense=H).
        _cell(ws, r, 8,
              f'=SUMIFS({exp}!$J:$J,{exp}!$E:$E,$A{r},'
              f'{exp}!$H:$H,"{EXPENSE_TYPE_PROFESSIONAL}"{exp_s})',
              fmt=INR, border=True)
        # Reimbursement Expenses: reimbursement-sheet outlays (Amount=I)
        # by the BOOKED Employee CC column (E, v0.3.109) — the cost
        # follows the employee's own cost centre, not the client's.
        _cell(ws, r, 9, f"=SUMIFS({reimb}!$I:$I,{reimb}!$E:$E,$A{r})",
              fmt=INR, border=True)
        # Provisions: outstanding provision remaining (Remaining=G, CC=C).
        _cell(ws, r, 10, f"=SUMIFS({prov}!$G:$G,{prov}!$C:$C,$A{r})",
              fmt=INR, border=True)
        # Total Direct Costs = Salary(both) + Professional Fees +
        # Reimbursement Expenses + Provisions.
        _cell(ws, r, 11, f"=F{r}+G{r}+H{r}+I{r}+J{r}", fmt=INR, border=True)
        # Gross Profit = Total Income − Total Direct Costs.
        _cell(ws, r, 12, f"=E{r}-K{r}", fmt=INR, border=True)
        # Gross Profit % = Gross ÷ Revenue (sales income only).
        _cell(ws, r, 13, f"=IF(C{r}=0,0,L{r}/C{r})",
              fmt=PCT, border=True, align=_CENTER)
        # Office Overhead (allocated): Overhead-type rows of the Salary sheet.
        _cell(ws, r, 14,
              f'=SUMIFS({lab}!$I:$I,{lab}!$C:$C,$A{r},'
              f'{lab}!$J:$J,"Overhead")',
              fmt=INR, border=True)
        # Indirect Expenses: Expenses with Type = Indirect Expense.
        _cell(ws, r, 15,
              f'=SUMIFS({exp}!$J:$J,{exp}!$E:$E,$A{r},'
              f'{exp}!$H:$H,"{EXPENSE_TYPE_INDIRECT}"{exp_s})',
              fmt=INR, border=True)
        # Net Profit = Gross − Office Overhead − Indirect Expenses.
        _cell(ws, r, 16, f"=L{r}-N{r}-O{r}", fmt=INR, border=True)
        # Net Profit % = Net ÷ Revenue (sales income only).
        _cell(ws, r, 17, f"=IF(C{r}=0,0,P{r}/C{r})",
              fmt=PCT, border=True, align=_CENTER)

    row_of: dict[str, int] = {}
    r = first
    for c in partners:
        write_row(r, c["code"], c["name"], "partner")
        row_of[c["code"]] = r
        r += 1
    for c in office:
        write_row(r, c["code"], c["name"], "office")
        row_of[c["code"]] = r
        r += 1
    if has_unassigned:
        write_row(r, "Unassigned", "Unassigned / to be mapped", "unassigned")
        row_of["Unassigned"] = r
        r += 1

    # Total row. Percentage columns (13 = Gross %, 17 = Net %) recompute on
    # the firm totals; every other column is a plain SUM.
    pct_cols = {13, 17}
    _cell(ws, total_row, 1, "", fill=_TOTAL_FILL)
    _cell(ws, total_row, 2, "TOTAL", font=_BOLD, fill=_TOTAL_FILL, border=True)
    for col in range(3, 18):
        L = get_column_letter(col)
        if col in pct_cols:
            num = "L" if col == 13 else "P"
            formula = (f"=IF(C{total_row}=0,0,{num}{total_row}/C{total_row})")
        else:
            formula = f"=SUM({L}{first}:{L}{last})"
        _cell(ws, total_row, col, formula, font=_BOLD, fill=_TOTAL_FILL,
              fmt=PCT if col in pct_cols else INR, border=True)

    # Filtered-subtotal row on top + AutoFilter over the CC rows (amount
    # columns only; the two % columns are excluded).
    amount_cols = [c for c in range(3, 18) if c not in pct_cols]
    _subtotal_filter(ws, hrow, first, last, amount_cols, last_col=17)
    ws.freeze_panes = f"A{first}"
    return {"first": first, "last": last, "total_row": total_row,
            "col_revenue": "C", "col_reimb_income": "D",
            "col_total_income": "E", "col_gross": "L", "col_profit": "P",
            "col_margin": "Q", "row_of": row_of}


# --- Partner – Manager P&L (matrix layout) ----------------------------------

def _build_pm_matrix(data: MISData, lbl: dict,
                     prior: "Window | None" = None,
                     cur_label: str | None = None):
    """Build the partner-manager column structure for the P&L matrix.

    Returns a list of partner blocks, where each block is:
        (cc_code, cc_name, columns)

    ``columns`` is a list of ``(label, role, filter)`` tuples. ``role`` is
    one of:

    * ``"self"``      — the partner's own direct work (no manager named).
      ``filter`` is ``"(unassigned)"`` — the value the data sheets write
      in the Manager column for an unassigned split.
    * ``"manager"``   — a named (active) manager sub-column. ``filter`` is
      the manager's code.
    * ``"subtotal"``  — the per-partner Total column (only present when the
      partner has at least one manager sub-column). ``filter`` is ``None``.
    * ``"partner_total"`` — a single partner-total column used when the partner
      has NO active managers (none assigned, or all deactivated in the master).
      ``filter`` is ``None`` — the cell SUMIFS at partner level with no
      manager criterion, so the partner's whole figure (including any
      deactivated manager's work) rolls up into this one column.
    * ``"prior_total"`` — the financial-year months BEFORE the selected
      period, as ONE cumulative column (v0.3.121; v0.3.120 broke it out per
      month, which made the sheet unreadably wide). ``filter`` is ``None``.

    A manager whose master row is deactivated is already folded to ``None`` on
    the facts by the calc engine, so it never appears as a "manager" column
    here; its work shows up under the partner's "self"/"partner_total" column.

    With an active *prior* window every block therefore ends with a single
    column — a June MIS reads "Jun-26 | Apr-26 to May-26".
    """
    cc_active_map = {c["id"]: c for c in lbl["cc_active"]}

    pairs: set[tuple[int, int | None]] = set()
    for f in (data.revenue_facts + data.expense_facts + data.labour_facts
              + data.provision_facts):
        cc_id = f.get("cost_centre_id")
        if cc_id is None:
            continue
        mgr_id = f.get("manager_id")  # labour / provision facts have none
        pairs.add((cc_id, mgr_id))

    by_partner: dict[int, list] = {}
    for cc_id, mgr_id in pairs:
        cc = cc_active_map.get(cc_id)
        if cc is None or cc["cc_type"] != "partner":
            continue
        by_partner.setdefault(cc_id, []).append(mgr_id)

    # A partner with previous-window activity but nothing in the CURRENT
    # period must still get a block (a collapsed 0 current column + the
    # month columns) — otherwise the MIS previous-months total would
    # silently undercount the firm's year to date (v0.3.102).
    prior_active = prior is not None and prior.active
    if prior_active:
        fy_data = prior.data
        for f in (fy_data.revenue_facts + fy_data.expense_facts
                  + fy_data.labour_facts + fy_data.provision_facts):
            cc_id = f.get("cost_centre_id")
            if cc_id is None or cc_id in by_partner:
                continue
            cc = cc_active_map.get(cc_id)
            if cc is not None and cc["cc_type"] == "partner":
                by_partner[cc_id] = [None]

    result = []
    for cc_id in sorted(by_partner.keys(),
                        key=lambda i: cc_active_map[i]["code"]):
        cc = cc_active_map[cc_id]
        mgr_ids = set(by_partner[cc_id])
        has_self = None in mgr_ids
        mgr_ids.discard(None)
        named = sorted(mgr_ids, key=lambda m: lbl["mgr"].get(m, "?") or "?")
        columns: list[tuple[str, str, str | None]] = []
        # Total-column labels carry the PERIOD (v0.3.103, operator ask):
        # the partner's name already sits in the merged super-header, so
        # "Jul-26" (or "Nov-26 to Feb-27") beside the cumulative label
        # reads much clearer than the partner's initials / "Total".
        total_lbl = cur_label or "Total"
        if named:
            # Partner delegates to at least one active manager — break the
            # partner's P&L down by manager, with a "Self" column for the
            # partner's own direct work and a per-partner subtotal.
            if has_self:
                columns.append((cc["code"], "self", "(unassigned)"))
            for mgr_id in named:
                columns.append((lbl["mgr"].get(mgr_id, "?"), "manager",
                                lbl["mgr"].get(mgr_id)))
            columns.append((total_lbl, "subtotal", None))
        else:
            # No active managers under this partner — show a single
            # partner-total column instead of an identical Self + Total pair.
            columns.append((cur_label or cc["code"], "partner_total", None))
        if prior_active:
            # The financial year so far, before the selected period, as one
            # cumulative column (v0.3.121) — partner-level SUMIFS over the
            # " (FY)" data sheets.
            columns.append((prior.label, "prior_total", None))
        result.append((cc["code"], cc["name"], columns))
    return result


def _sheet_partner_manager(wb: Workbook, data: MISData, lbl: dict,
                           prior: "Window | None" = None) -> None:
    """The headline P&L sheet — partner super-headers, manager sub-columns,
    formula-driven from the Revenue / Expenses / Labour data sheets.

    Each partner block ends with ONE cumulative column for the financial
    year so far — the FY months BEFORE the selected period (*prior*), so
    a June MIS reads "Jun-26 | Apr-26 to May-26". The year-to-date
    against last year lives on "Partner-Manager P&L (Cmp)"."""
    title = "Partner – Manager Profitability"
    ws = wb.create_sheet("Partner-Manager P&L")
    ws.sheet_view.showGridLines = False

    prior_active = prior is not None and prior.active
    prior_sfx = prior.suffix if prior_active else FYS
    cur_lbl = periods_label(data.options.periods)
    matrix = _build_pm_matrix(data, lbl, prior=prior, cur_label=cur_lbl)
    if not matrix:
        # Nothing to show — still draw a heading so the sheet isn't blank.
        _cell(ws, 1, 1, title, font=_TITLE)
        _cell(ws, 3, 1,
              "No data yet — once you import sales / expenses / salary the "
              "full matrix will appear here.", font=_SUB)
        return

    # ---- Title + subtitle ----
    _cell(ws, 1, 1, title, font=_TITLE)
    _cell(ws, 2, 1,
          "Period(s): " + periods_label(data.options.periods),
          font=_SUB)

    # ---- Column layout ----
    # Col 1 = Particulars. Then each partner block contributes its columns
    # (leaves + the selected period's total + the FY-so-far column). The
    # final block is the firm-wide "MIS Total", which mirrors a partner
    # block's shape: the selected period, then the same FY-so-far column.
    block_starts: list[int] = []
    cols_per_block: list[int] = []
    col_idx = 2
    for cc_code, cc_name, managers in matrix:
        block_starts.append(col_idx)
        cols_per_block.append(len(managers))
        col_idx += len(managers)
    # (label, role, filter) for the MIS Total block, same tuple shape as a
    # partner block's columns so both render through the same code.
    mis_cols: list[tuple[str, str, str | None]] = [
        (cur_lbl or "Total", "current", None)]
    if prior_active:
        mis_cols.append((prior.label, "prior_total", None))
    mis_start = col_idx                  # first cell = the selected period
    last_col = mis_start + len(mis_cols) - 1

    # Set column widths (total columns are wider — they carry a period-range
    # label like "Apr-26 to May-26").
    ws.column_dimensions["A"].width = 30
    for col in range(2, last_col + 1):
        ws.column_dimensions[get_column_letter(col)].width = 14
    for (cc_code, cc_name, managers), start in zip(matrix, block_starts):
        for off, (_lbl2, _role2, _f2) in enumerate(managers):
            if _role2 in ("prior_total", "subtotal", "partner_total"):
                ws.column_dimensions[
                    get_column_letter(start + off)].width = 17
    for off in range(len(mis_cols)):
        ws.column_dimensions[get_column_letter(mis_start + off)].width = 17

    # ---- Header rows ----
    hdr_partner_row = 4
    hdr_mgr_row = 5
    body_start = 7

    def _super_header(name: str, start: int, count: int) -> None:
        end = start + count - 1
        if end > start:   # a collapsed single-column block needs no merge
            ws.merge_cells(start_row=hdr_partner_row, start_column=start,
                           end_row=hdr_partner_row, end_column=end)
        _cell(ws, hdr_partner_row, start, name, font=_HEAD,
              fill=_HEAD_FILL, align=_CENTER, border=True)
        for col in range(start, end + 1):
            ws.cell(row=hdr_partner_row, column=col).fill = _HEAD_FILL
            ws.cell(row=hdr_partner_row, column=col).border = _BORDER

    # Partner super-header (row 4): merged across each block's columns.
    for (cc_code, cc_name, managers), start, count in zip(
            matrix, block_starts, cols_per_block):
        _super_header(cc_name, start, count)
    _super_header("MIS Total", mis_start, len(mis_cols))
    _cell(ws, hdr_partner_row, 1, "", fill=_HEAD_FILL)

    # Sub-header (row 5): one cell per column — manager codes, the period
    # totals and the previous-month labels.
    _cell(ws, hdr_mgr_row, 1, "Particulars", font=_HEAD, fill=_SUBHEAD_FILL,
          align=_CENTER, border=True)
    for (cc_code, cc_name, managers), start in zip(matrix, block_starts):
        for offset, (label, _role, _filter) in enumerate(managers):
            _cell(ws, hdr_mgr_row, start + offset, label, font=_HEAD,
                  fill=_SUBHEAD_FILL, align=_CENTER, border=True)
    for offset, (label, _role, _filter) in enumerate(mis_cols):
        _cell(ws, hdr_mgr_row, mis_start + offset, label, font=_HEAD,
              fill=_SUBHEAD_FILL, align=_CENTER, border=True)

    # ---- P&L lines ----
    # Revenue / Expenses hold both period windows on one sheet (v0.3.118),
    # so the selected-period cells carry a Scope criterion; Salary /
    # Reimbursements / Provisions still have per-window sheets.
    rev, exp = _data_sheet_q("Revenue"), _data_sheet_q("Expenses")
    lab = _q("Salary")
    reimb = _q("Reimbursements")
    prov = _q("Provisions")
    rev_s = _scope_crit(rev, SCOPE_COL_REV)
    exp_s = _scope_crit(exp, SCOPE_COL_EXP)

    # Returns the SUMIFS formula for a (partner, manager) cell on a given
    # data sheet's amount column.
    def sumifs(sheet_q, amount_col, cc_code, mgr_filter, extra=None,
               cc_col="D", mgr_col="E", scope=""):
        # Column layout on the data sheets (v0.3.69, + Scope in v0.3.118):
        #   Revenue:  A=Date B=VoucherNo C=Entity D=CostCentre E=Manager
        #             F=Service G=Client H=Amount I=Category J=Period
        #             K=Scope
        #   Expenses: A=Date B=VoucherNo C=InvoiceNo D=Entity E=CostCentre
        #             F=Manager G=Service H=TypeOfExpense I=Client
        #             J=Amount K=Description L=Period M=Scope
        parts = [
            f"{sheet_q}!${amount_col}:${amount_col}",
            f"{sheet_q}!${cc_col}:${cc_col}", f'"{cc_code}"',
        ]
        if mgr_filter is not None:
            parts.append(f"{sheet_q}!${mgr_col}:${mgr_col}")
            parts.append(f'"{mgr_filter}"')
        for col, value in (extra or []):
            parts.append(f"{sheet_q}!${col}:${col}")
            parts.append(f'"{value}"')
        return "=SUMIFS(" + ",".join(parts) + scope + ")"

    def exp_sumifs(cc_code, mgr_filter, type_label):
        # Expense cells filter on the Type of Expense column (H) so the
        # P&L can split Professional Fees (direct) from Indirect
        # Expenses (shown under the overhead block).
        return sumifs(exp, "J", cc_code, mgr_filter,
                      extra=[("H", type_label)], cc_col="E", mgr_col="F",
                      scope=exp_s)

    def labour_sumifs(cc_code, billable=None, mgr_filter=None):
        # Labour now carries the employee's manager (Salary sheet col L,
        # v0.3.85), so salary breaks down by manager just like revenue /
        # expenses. Filter by Type="Salary" so the Overhead facts don't
        # double-count (they have their own row). ``billable`` ("Yes"/"No",
        # col K) splits client-billable from non-billable / office time;
        # ``mgr_filter`` (manager code, or "(unassigned)" for the partner's
        # own column) places the cost in the right manager sub-column.
        extra = f',{lab}!$K:$K,"{billable}"' if billable is not None else ""
        if mgr_filter is not None:
            extra += f',{lab}!$L:$L,"{mgr_filter}"'
        return (f'=SUMIFS({lab}!$I:$I,{lab}!$C:$C,"{cc_code}",'
                f'{lab}!$J:$J,"Salary"{extra})')

    # Row layout. Each entry: (label, kind)
    # kinds:
    #   "sales"     -> Revenue, Category = Income
    #   "reimb"     -> Revenue, Category IN (Reimbursement, OPE)
    #   "salary"    -> Labour amount (partner-level only)
    #   "expense"   -> Expenses, Type = Professional Fees (DIRECT cost)
    #   "reimb_exp" -> Reimbursements sheet (partner-level only)
    #   "income_sum"-> SUM of sales + reimb rows
    #   "cost_sum"  -> SUM of salary + expense + reimb_exp rows
    #   "gross"     -> income_sum - cost_sum
    #   "gross_pct" -> gross / income
    #   "overhead"  -> Salary sheet, Type = Overhead (per-employee share
    #                  of office indirect costs — see Employee Register)
    #   "indirect"  -> Expenses, Type = Indirect Expense (the partner's
    #                  own indirect costs, NOT the office pool)
    #   "net"       -> gross - overhead - indirect
    #   "net_pct"   -> net / income
    # We track by row number for cross-referencing.
    plan = [
        ("Sales (Income)", "sales"),
        ("Reimbursement & OPE", "reimb"),
        ("Total Income", "income_sum"),
        ("", "blank"),
        ("Salary (billable)", "salary"),
        ("Salary (non-billable)", "salary_nonbill"),
        ("Professional Fees", "expense"),
        ("Reimbursement Expenses", "reimb_exp"),
        ("Provisions", "provisions"),
        ("Total Direct Costs", "cost_sum"),
        ("", "blank"),
        ("Gross Profit", "gross"),
        ("Gross Profit %", "gross_pct"),
        ("", "blank"),
        ("Office Overhead (allocated)", "overhead"),
        ("Indirect Expenses", "indirect"),
        ("Net Profit", "net"),
        ("Net Profit %", "net_pct"),
    ]

    r = body_start
    rows_by_kind: dict[str, int] = {}
    for label, kind in plan:
        if kind == "blank":
            r += 1
            continue
        rows_by_kind[kind] = r
        is_total = kind in ("income_sum", "cost_sum", "gross", "gross_pct",
                            "net", "net_pct")
        font = _BOLD if is_total else _NORMAL
        fill = _TOTAL_FILL if is_total else None
        _cell(ws, r, 1, label, font=font, fill=fill, border=True)

        for (cc_code, cc_name, managers), start in zip(matrix, block_starts):
            for offset, (mgr_label, role, mgr_filter) in enumerate(managers):
                col = start + offset
                is_total_col = (role == "subtotal")
                is_partner_total = (role == "partner_total")
                cell_fmt = PCT if kind.endswith("_pct") else INR
                L = get_column_letter(col)
                prior_leaf = (_pm_leaf_formula(kind, cc_code, prior_sfx)
                              if role == "prior_total" else None)
                if prior_leaf is not None:
                    # The FY-so-far column: partner-level SUMIFS against
                    # the " (FY)" data sheets. Row-arithmetic kinds
                    # (totals / %s) fall through to the generic
                    # same-column branches below, which are column-local
                    # and already correct here.
                    formula = prior_leaf
                elif is_total_col and kind == "gross_pct":
                    # Recompute the ratio at the partner level — summing
                    # percentages doesn't make sense. Denominator is SALES
                    # (income only), not Total Income (v0.3.92).
                    inc_r = rows_by_kind["sales"]
                    gross_r = rows_by_kind["gross"]
                    L = get_column_letter(col)
                    formula = (f"=IF({L}{inc_r}=0,0,"
                               f"{L}{gross_r}/{L}{inc_r})")
                elif (is_total_col or is_partner_total) and kind == "overhead":
                    # Allocated office overhead is partner-level, not
                    # manager-level — pull from the Salary sheet via
                    # SUMIFS on Type="Overhead" so it stays formula-
                    # driven (v0.3.67). Pre-v0.3.67 baked the computed
                    # number directly into the cell. A collapsed
                    # partner-total column carries it the same way.
                    formula = (f'=SUMIFS({lab}!$I:$I,{lab}!$C:$C,'
                               f'"{cc_code}",{lab}!$J:$J,"Overhead")')
                elif is_total_col and kind == "net":
                    gross_r = rows_by_kind["gross"]
                    overhead_r = rows_by_kind["overhead"]
                    indirect_r = rows_by_kind["indirect"]
                    L = get_column_letter(col)
                    formula = (f"={L}{gross_r}-{L}{overhead_r}"
                               f"-{L}{indirect_r}")
                elif is_total_col and kind == "net_pct":
                    # Denominator is SALES (income only), not Total Income.
                    inc_r = rows_by_kind["sales"]
                    net_r = rows_by_kind["net"]
                    L = get_column_letter(col)
                    formula = (f"=IF({L}{inc_r}=0,0,"
                               f"{L}{net_r}/{L}{inc_r})")
                elif is_total_col:
                    # Plain SUM across the partner's leaf (Self + manager)
                    # columns — NOT the FY-cumulative column after the
                    # Total (that belongs to a different period window).
                    leaf_n = sum(1 for _l3, ro3, _f3 in managers
                                 if ro3 in ("self", "manager"))
                    from_col = start
                    to_col = start + leaf_n - 1
                    if from_col > to_col:
                        formula = 0
                    else:
                        formula = (f"=SUM({get_column_letter(from_col)}{r}:"
                                   f"{get_column_letter(to_col)}{r})")
                elif kind == "overhead":
                    # Office overhead is partner-level only — it doesn't
                    # break down by manager, so manager columns are blank
                    # (the partner Total column carries it).
                    formula = 0
                elif kind == "net":
                    # Net per manager = Gross − Overhead − Indirect. Overhead
                    # is 0 in manager columns, so this is Gross − Indirect;
                    # the office overhead is deducted only in the partner
                    # Total column (it isn't attributable to a manager).
                    g = rows_by_kind["gross"]
                    o = rows_by_kind["overhead"]
                    ind = rows_by_kind["indirect"]
                    L = get_column_letter(col)
                    formula = f"={L}{g}-{L}{o}-{L}{ind}"
                elif kind == "net_pct":
                    # Net % per manager = Net ÷ Sales (income only).
                    s = rows_by_kind["sales"]
                    nr = rows_by_kind["net"]
                    L = get_column_letter(col)
                    formula = f"=IF({L}{s}=0,0,{L}{nr}/{L}{s})"
                elif kind == "sales":
                    formula = sumifs(rev, "H", cc_code, mgr_filter,
                                     extra=[("I", "Income")], scope=rev_s)
                elif kind == "reimb":
                    # Reimbursement + OPE together — two SUMIFS summed.
                    f1 = sumifs(rev, "H", cc_code, mgr_filter,
                                extra=[("I", "Reimbursement")], scope=rev_s)
                    f2 = sumifs(rev, "H", cc_code, mgr_filter,
                                extra=[("I", "OPE")], scope=rev_s)
                    formula = "=" + f1[1:] + "+" + f2[1:]
                elif kind == "income_sum":
                    sales_r = rows_by_kind["sales"]
                    reimb_r = rows_by_kind["reimb"]
                    formula = (f"={get_column_letter(col)}{sales_r}+"
                               f"{get_column_letter(col)}{reimb_r}")
                elif kind == "salary":
                    # Billable labour, broken down by the employee's manager.
                    formula = labour_sumifs(cc_code, "Yes", mgr_filter)
                elif kind == "salary_nonbill":
                    # Non-billable / office-booked labour, by manager.
                    formula = labour_sumifs(cc_code, "No", mgr_filter)
                elif kind == "expense":
                    # Professional fees bought in — the partner's DIRECT
                    # expense (operator ask v0.3.69; was "all expenses").
                    formula = exp_sumifs(cc_code, mgr_filter,
                                         EXPENSE_TYPE_PROFESSIONAL)
                elif kind == "indirect":
                    formula = exp_sumifs(cc_code, mgr_filter,
                                         EXPENSE_TYPE_INDIRECT)
                elif kind == "reimb_exp":
                    # Reimbursement outlays land on the EMPLOYEE's own
                    # cost centre (Employee CC = E, v0.3.109) — partner-
                    # level ("Self" column) only. Keeps the P&L tying
                    # with the Cost Centre P&L's Reimbursement Expenses.
                    if offset == 0:
                        formula = (f'=SUMIFS({reimb}!$I:$I,'
                                   f'{reimb}!$E:$E,"{cc_code}")')
                    else:
                        formula = 0
                elif kind == "provisions":
                    # Outstanding provisions — CC-level (no manager), so the
                    # partner's "Self" column only. Provisions sheet:
                    # Remaining=G by CostCentre=C.
                    if offset == 0:
                        formula = (f'=SUMIFS({prov}!$G:$G,'
                                   f'{prov}!$C:$C,"{cc_code}")')
                    else:
                        formula = 0
                elif kind == "cost_sum":
                    sal_r = rows_by_kind["salary"]
                    snb_r = rows_by_kind["salary_nonbill"]
                    exp_r = rows_by_kind["expense"]
                    rexp_r = rows_by_kind["reimb_exp"]
                    prov_r = rows_by_kind["provisions"]
                    L = get_column_letter(col)
                    formula = (f"={L}{sal_r}+{L}{snb_r}+{L}{exp_r}"
                               f"+{L}{rexp_r}+{L}{prov_r}")
                elif kind == "gross":
                    inc_r = rows_by_kind["income_sum"]
                    cost_r = rows_by_kind["cost_sum"]
                    formula = (f"={get_column_letter(col)}{inc_r}-"
                               f"{get_column_letter(col)}{cost_r}")
                elif kind == "gross_pct":
                    inc_r = rows_by_kind["sales"]   # % on sales income only
                    gross_r = rows_by_kind["gross"]
                    L = get_column_letter(col)
                    formula = (f"=IF({L}{inc_r}=0,0,"
                               f"{L}{gross_r}/{L}{inc_r})")
                else:
                    formula = 0

                _cell(ws, r, col, formula, font=font, fill=fill, fmt=cell_fmt,
                      border=True)

        # ---- MIS Total block ----
        # One cell per column of the block: the selected period, then the
        # FY so far. Leaf kinds add up the partner blocks' matching
        # columns; arithmetic kinds (totals and the two %s) recompute
        # within their own column, so every % is a true ratio of that
        # column's own figures.
        def block_refs(role_match) -> list[str]:
            refs = []
            for (_cc4, _nm4, cols4), start in zip(matrix, block_starts):
                for off, (_l4, ro4, _f4) in enumerate(cols4):
                    if ro4 in role_match:
                        refs.append(f"{get_column_letter(start + off)}{r}")
                        break
            return refs

        for offset, (_mlbl, mrole, _mfilter) in enumerate(mis_cols):
            col = mis_start + offset
            L = get_column_letter(col)
            if mrole == "current":
                if kind == "gross_pct":
                    inc_r = rows_by_kind["sales"]     # % on sales (v0.3.92)
                    gross_r = rows_by_kind["gross"]
                    mis_formula = (f"=IF({L}{inc_r}=0,0,"
                                   f"{L}{gross_r}/{L}{inc_r})")
                elif kind == "net_pct":
                    inc_r = rows_by_kind["sales"]
                    net_r = rows_by_kind["net"]
                    mis_formula = (f"=IF({L}{inc_r}=0,0,"
                                   f"{L}{net_r}/{L}{inc_r})")
                else:
                    # Reference each block's subtotal / partner-total column
                    # — a block's LAST columns are the previous months.
                    refs = block_refs(("subtotal", "partner_total"))
                    mis_formula = "=" + "+".join(refs) if refs else 0
            else:
                mis_formula = _pm_col_arith(kind, L, rows_by_kind)
                if mis_formula is None:
                    # Leaf kind: add up the partner blocks' FY-so-far
                    # columns.
                    refs = block_refs(("prior_total",))
                    mis_formula = "=" + "+".join(refs) if refs else 0
            _cell(ws, r, col, mis_formula, font=_BOLD, fill=_TOTAL_FILL,
                  fmt=PCT if kind.endswith("_pct") else INR, border=True)

        r += 1

    ws.freeze_panes = f"B{body_start}"
    note = (
        "Salary (billable) = labour booked to this partner's clients; "
        "Salary (non-billable) = the partner's own staff time that isn't "
        "client-billable (office-booked / admin / unlogged), charged to "
        "their home cost centre. Office Overhead = (office indirect "
        "expenses + office-home staff salaries) ÷ partner-team employees, "
        "charged to each one's home cost centre (see Employee Register). "
        "Professional Fees / Indirect Expenses follow the Cost Center on "
        "the voucher split and the Type of Expense column. Office Overhead "
        "is partner-level (not split by manager), so a manager column's "
        "Net is before office overhead — the partner Total deducts it, and "
        "Gross/Net % are on sales income.")
    if prior_active:
        note += (
            f" The '{prior.label}' column after each partner's total is the "
            "CUMULATIVE figure for the financial-year months BEFORE the "
            "selected period (a selection starting in April shows the whole "
            "previous FY), read live from the Revenue / Expenses sheets' "
            "Scope = \"FY Prior\" rows and the ' (FY)' Salary / "
            "Reimbursements / Provisions sheets. It is context only and is "
            "NOT included in the partner Total or the MIS Total. Provisions "
            "carry forward, so that column shows the balance OUTSTANDING at "
            "the end of the window, never a sum of months. For the year to "
            "date against the same period last year, see the "
            "'Partner-Manager P&L (Cmp)' sheet.")
    _cell(ws, r + 1, 1, note, font=_SUB)


# --- Partner-level P&L formula builders (shared) ------------------------------
# Used by the Partner-Manager comparison sheet AND the previous-month columns
# on the main Partner-Manager P&L, so both stay formula-identical.

def _pm_leaf_formula(kind: str, cc_code: str, sfx: str) -> str | None:
    """Partner-level SUMIFS for *kind* against the ``sfx`` data sheets —
    ``""`` (the selected period), :data:`FYS`, :data:`YTD` or :data:`LY`.
    Returns ``None`` for row-arithmetic kinds (sums / %s).

    Revenue / Expenses are ONE sheet for the first two since v0.3.118:
    those windows live together and are told apart by the Scope
    criterion. The year-on-year windows name their own sheets. Salary /
    Reimbursements / Provisions have a sheet per window throughout.
    """
    rev, exp = _data_sheet_q("Revenue", sfx), _data_sheet_q("Expenses", sfx)
    lab = _q("Salary" + sfx)
    reimb, prov = _q("Reimbursements" + sfx), _q("Provisions" + sfx)
    rev_s = _scope_crit(rev, SCOPE_COL_REV, sfx)
    exp_s = _scope_crit(exp, SCOPE_COL_EXP, sfx)
    if kind == "sales":
        return (f'=SUMIFS({rev}!$H:$H,{rev}!$D:$D,"{cc_code}",'
                f'{rev}!$I:$I,"Income"{rev_s})')
    if kind == "reimb":
        f1 = (f'SUMIFS({rev}!$H:$H,{rev}!$D:$D,"{cc_code}",'
              f'{rev}!$I:$I,"Reimbursement"{rev_s})')
        f2 = (f'SUMIFS({rev}!$H:$H,{rev}!$D:$D,"{cc_code}",'
              f'{rev}!$I:$I,"OPE"{rev_s})')
        return "=" + f1 + "+" + f2
    if kind == "salary":
        return (f'=SUMIFS({lab}!$I:$I,{lab}!$C:$C,"{cc_code}",'
                f'{lab}!$J:$J,"Salary",{lab}!$K:$K,"Yes")')
    if kind == "salary_nonbill":
        return (f'=SUMIFS({lab}!$I:$I,{lab}!$C:$C,"{cc_code}",'
                f'{lab}!$J:$J,"Salary",{lab}!$K:$K,"No")')
    if kind == "expense":
        return (f'=SUMIFS({exp}!$J:$J,{exp}!$E:$E,"{cc_code}",'
                f'{exp}!$H:$H,"{EXPENSE_TYPE_PROFESSIONAL}"{exp_s})')
    if kind == "indirect":
        return (f'=SUMIFS({exp}!$J:$J,{exp}!$E:$E,"{cc_code}",'
                f'{exp}!$H:$H,"{EXPENSE_TYPE_INDIRECT}"{exp_s})')
    if kind == "reimb_exp":
        # Booked by the Employee CC column (E, v0.3.109) — the same
        # criterion on every window's edition of the sheet.
        return f'=SUMIFS({reimb}!$I:$I,{reimb}!$E:$E,"{cc_code}")'
    if kind == "provisions":
        return f'=SUMIFS({prov}!$G:$G,{prov}!$C:$C,"{cc_code}")'
    if kind == "overhead":
        return (f'=SUMIFS({lab}!$I:$I,{lab}!$C:$C,"{cc_code}",'
                f'{lab}!$J:$J,"Overhead")')
    return None


def _pm_col_arith(kind: str, L: str, rows_by_kind: dict) -> str | None:
    """Row-arithmetic P&L kinds (totals and %s), computed within one
    column letter *L* from that column's own leaf rows."""
    if kind == "income_sum":
        return f"={L}{rows_by_kind['sales']}+{L}{rows_by_kind['reimb']}"
    if kind == "cost_sum":
        return (f"={L}{rows_by_kind['salary']}"
                f"+{L}{rows_by_kind['salary_nonbill']}"
                f"+{L}{rows_by_kind['expense']}"
                f"+{L}{rows_by_kind['reimb_exp']}"
                f"+{L}{rows_by_kind['provisions']}")
    if kind == "gross":
        return (f"={L}{rows_by_kind['income_sum']}"
                f"-{L}{rows_by_kind['cost_sum']}")
    if kind == "gross_pct":
        s, g = rows_by_kind["sales"], rows_by_kind["gross"]
        return f"=IF({L}{s}=0,0,{L}{g}/{L}{s})"
    if kind == "net":
        return (f"={L}{rows_by_kind['gross']}"
                f"-{L}{rows_by_kind['overhead']}"
                f"-{L}{rows_by_kind['indirect']}")
    if kind == "net_pct":
        s, n = rows_by_kind["sales"], rows_by_kind["net"]
        return f"=IF({L}{s}=0,0,{L}{n}/{L}{s})"
    return None


# --- Partner-Manager P&L: current vs comparison ------------------------------

def _sheet_pm_comparison(wb: Workbook, data: MISData, lbl: dict,
                         windows: "Windows") -> None:
    """"Partner-Manager P&L (Cmp)" — the SAME P&L lines as the main
    Partner-Manager sheet, but per partner three columns side by side:
    **year to date | the same span last year | Δ**.

    v0.3.121 turned this into a fixed year-on-year view (operator ask).
    It used to compare the selected period against months the operator
    ticked under "Compare with"; that selection is gone, and the two
    windows are now derived — a Jun-26 MIS reads "Apr-26 to Jun-26" vs
    "Apr-25 to Jun-25".

    Partner-level (no manager sub-columns — the manager breakdown for the
    selected period lives on the main Partner-Manager P&L). Every cell is
    a live formula: the year to date SUMIFS the " (YTD)" data sheets,
    last year the " (LY)" ones, Δ = year to date − last year.
    """
    ws = wb.create_sheet("Partner-Manager P&L (Cmp)")
    ws.sheet_view.showGridLines = False

    _cell(ws, 1, 1, "Partner – Manager P&L — Year to date vs last year",
          font=_TITLE)
    cur_lbl = windows.ytd.label
    cmp_lbl = windows.last_year.label
    _cell(ws, 2, 1,
          f"Financial year to date: {cur_lbl}    vs    the same period last "
          f"year: {cmp_lbl}.  Reporting period: "
          f"{periods_label(windows.selected)}.  Partner-level; the manager "
          "breakdown for the reporting period is on the Partner-Manager "
          "P&L sheet.",
          font=_SUB)

    partners = [c for c in lbl["cc_active"] if c["cc_type"] == "partner"]
    if not partners:
        _cell(ws, 4, 1, "No partner cost centres are active.", font=_SUB)
        return

    hdr_partner_row = 4
    hdr_sub_row = 5
    body_start = 7
    n_cols_per_block = 3          # Current | Comparison | Δ

    ws.column_dimensions["A"].width = 30
    block_starts = [2 + i * n_cols_per_block for i in range(len(partners))]
    mis_start = 2 + len(partners) * n_cols_per_block
    last_col = mis_start + n_cols_per_block - 1
    for col in range(2, last_col + 1):
        ws.column_dimensions[get_column_letter(col)].width = 17

    # Partner super-headers + the MIS Total block.
    blocks = ([(c["name"], start) for c, start in zip(partners, block_starts)]
              + [("MIS Total", mis_start)])
    for name, start in blocks:
        end = start + n_cols_per_block - 1
        ws.merge_cells(start_row=hdr_partner_row, start_column=start,
                       end_row=hdr_partner_row, end_column=end)
        _cell(ws, hdr_partner_row, start, name, font=_HEAD, fill=_HEAD_FILL,
              align=_CENTER, border=True)
        for col in range(start, end + 1):
            ws.cell(row=hdr_partner_row, column=col).fill = _HEAD_FILL
            ws.cell(row=hdr_partner_row, column=col).border = _BORDER
        for off, label in enumerate((cur_lbl, cmp_lbl, "Δ")):
            _cell(ws, hdr_sub_row, start + off, label, font=_HEAD,
                  fill=_SUBHEAD_FILL, align=_CENTER, border=True)
    _cell(ws, hdr_partner_row, 1, "", fill=_HEAD_FILL)
    _cell(ws, hdr_sub_row, 1, "Particulars", font=_HEAD, fill=_SUBHEAD_FILL,
          align=_CENTER, border=True)

    # Shared partner-level builders (also used by the FY-so-far column on
    # the main sheet), so both sheets stay formula-identical.
    leaf_formula = _pm_leaf_formula

    # Same line plan as the main Partner-Manager P&L.
    plan = [
        ("Sales (Income)", "sales"),
        ("Reimbursement & OPE", "reimb"),
        ("Total Income", "income_sum"),
        ("", "blank"),
        ("Salary (billable)", "salary"),
        ("Salary (non-billable)", "salary_nonbill"),
        ("Professional Fees", "expense"),
        ("Reimbursement Expenses", "reimb_exp"),
        ("Provisions", "provisions"),
        ("Total Direct Costs", "cost_sum"),
        ("", "blank"),
        ("Gross Profit", "gross"),
        ("Gross Profit %", "gross_pct"),
        ("", "blank"),
        ("Office Overhead (allocated)", "overhead"),
        ("Indirect Expenses", "indirect"),
        ("Net Profit", "net"),
        ("Net Profit %", "net_pct"),
    ]

    arith_formula = _pm_col_arith

    r = body_start
    rows_by_kind: dict[str, int] = {}
    for label, kind in plan:
        if kind == "blank":
            r += 1
            continue
        rows_by_kind[kind] = r
        is_total = kind in ("income_sum", "cost_sum", "gross", "gross_pct",
                            "net", "net_pct")
        font = _BOLD if is_total else _NORMAL
        fill = _TOTAL_FILL if is_total else None
        cell_fmt = PCT if kind.endswith("_pct") else INR
        _cell(ws, r, 1, label, font=font, fill=fill, border=True)

        for c, start in zip(partners, block_starts):
            for off, sfx in ((0, YTD), (1, LY)):
                L = get_column_letter(start + off)
                formula = (leaf_formula(kind, c["code"], sfx)
                           or arith_formula(kind, L, rows_by_kind))
                _cell(ws, r, start + off, formula, font=font, fill=fill,
                      fmt=cell_fmt, border=True)
            # Δ = year to date − last year (percentage-point diff on %s).
            curL = get_column_letter(start)
            cmpL = get_column_letter(start + 1)
            _cell(ws, r, start + 2, f"={curL}{r}-{cmpL}{r}", font=font,
                  fill=fill, fmt=cell_fmt, border=True)

        # MIS Total block: each side sums the partner columns (leaf
        # kinds) or recomputes from this block's own rows (arithmetic
        # kinds, incl. the two %s); Δ = year to date − last year.
        for off in (0, 1):
            L = get_column_letter(mis_start + off)
            if leaf_formula(kind, "X", YTD) is not None:
                refs = [f"{get_column_letter(start + off)}{r}"
                        for start in block_starts]
                formula = "=" + "+".join(refs)
            else:
                formula = arith_formula(kind, L, rows_by_kind)
            _cell(ws, r, mis_start + off, formula, font=_BOLD,
                  fill=_TOTAL_FILL, fmt=cell_fmt, border=True)
        curL = get_column_letter(mis_start)
        cmpL = get_column_letter(mis_start + 1)
        _cell(ws, r, mis_start + 2, f"={curL}{r}-{cmpL}{r}", font=_BOLD,
              fill=_TOTAL_FILL, fmt=cell_fmt, border=True)
        r += 1

    ws.freeze_panes = f"B{body_start}"
    _cell(ws, r + 1, 1,
          "The year-to-date columns read the ' (YTD)' data sheets, last "
          "year's the ' (LY)' ones — both cover the financial year from "
          "April through the reporting month, one year apart. "
          "Δ = year to date − last year (percentage-point "
          "difference on the % rows). Provisions show the balance "
          "outstanding at the end of each window. Office Overhead and Net "
          "follow the same definitions as the Partner-Manager P&L sheet.",
          font=_SUB)


# --- Entity & Service --------------------------------------------------------

def _sheet_entity(wb: Workbook, data: MISData, lbl: dict) -> None:
    # Entity is column C on Revenue, D on Expenses (v0.3.69 layout).
    _simple_summary(wb, "Entity P&L", "Entity-wise Profitability",
                    "Entity", data.entities, "ent", lbl, key="name",
                    sumcol_rev="C", sumcol_exp="D", prov_col="B")


def _sheet_service(wb: Workbook, data: MISData, lbl: dict) -> None:
    # Service is column F on Revenue, G on Expenses (v0.3.69 layout).
    _simple_summary(wb, "Service MIS", "Service-wise Revenue & Cost",
                    "Service", data.services, "svc", lbl, key="name",
                    sumcol_rev="F", sumcol_exp="G")


def _sheet_client_billing(wb: Workbook, data: MISData, lbl: dict,
                          prior: "Window | None" = None) -> None:
    """Client × period billing matrix.

    One row per client (canonical name from the master, or ``(unmapped)``
    when the voucher's party didn't resolve). Columns: Client | Grand
    Total | one per PREVIOUS month | one per selected period. EVERY month
    cell is a live SUMIFS by Client + Period label over a revenue
    register (v0.3.117; they were baked values) — the "Revenue" sheet,
    which since v0.3.118 carries the prior-FY months too, or
    "Revenue (Cmp)" for an operator-chosen comparison window (v0.3.120)
    — so a sales row edited or added post-generation flows straight
    through. Clients billed only in the previous window still get a row
    (0 current), so the history never undercounts. The Grand Total sums
    ONLY the selected period's columns. Credit / Debit Notes flow through
    with signs intact. Sorted by Grand Total descending.
    """
    if not data.options.periods:
        return
    periods = sorted(data.options.periods)
    has_fy = bool(prior is not None and prior.active)
    fy_months = sorted(prior.periods) if has_fy else []
    fy_label = prior.label if has_fy else None
    fy_data = prior.data if has_fy else None
    prior_rev = _data_sheet_q("Revenue", prior.suffix) if has_fy else None

    # Group by resolved client; unresolved parties group by their raw
    # party name (one row per distinct vendor/client as Tally spells it)
    # instead of all lumping into a single "(unmapped)" bucket.
    agg: dict = {}
    names: dict = {}

    def client_key(f):
        cid = f.get("client_id")
        if cid is not None:
            key = ("c", cid)
            names[key] = lbl["cli"].get(cid, "(unmapped)")
        else:
            party = (f.get("party_name") or "").strip()
            key = ("r", norm(party)) if party else ("u",)
            names[key] = party or "(unmapped)"
        return key

    for f in data.revenue_facts:
        key = client_key(f)
        bucket = agg.setdefault(key, {p: 0.0 for p in periods})
        bucket[f["period"]] = bucket.get(f["period"], 0.0) + float(f["amount"])
    # Clients billed only in the FY-prior window get a zero current row —
    # their cumulative figure must still show.
    if has_fy:
        for f in fy_data.revenue_facts:
            key = client_key(f)
            agg.setdefault(key, {p: 0.0 for p in periods})

    if not agg:
        return

    # Build sortable list: (name, total, [per-period amounts])
    rows = []
    for key, per_period in agg.items():
        amounts = [round(per_period.get(p, 0.0), 2) for p in periods]
        total = round(sum(amounts), 2)
        rows.append((names[key], total, amounts))
    rows.sort(key=lambda r: -r[1])

    # Layout: A Client, B Grand Total, then one column per FY month
    # before the selected period, then one column per selected period.
    fy0 = 3                              # first prior-FY month column
    p0 = 3 + len(fy_months)              # first current-period column

    ws = wb.create_sheet("Client Billing")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 16
    for i in range(len(fy_months) + len(periods)):
        ws.column_dimensions[get_column_letter(fy0 + i)].width = 13

    _cell(ws, 1, 1, "Client-wise Billing", font=_TITLE)
    _cell(ws, 2, 1, "Period(s): " + periods_label(periods)
          + (f".  Columns {fy_label} show the earlier months of this "
             "financial year, month by month; the Grand Total sums only "
             "the selected period's columns." if has_fy else ""),
          font=_SUB)
    hrow = 4
    headers = (["Client", "Grand Total"]
               + [_month_short(p) for p in fy_months]
               + [_month_short(p) for p in periods])
    _header_row(ws, hrow, headers)

    body_start = hrow + 1
    r = body_start
    first_period_col = get_column_letter(p0)
    last_period_col = get_column_letter(p0 + len(periods) - 1)
    rev = _q("Revenue")
    for name, _total, _amounts in rows:
        _cell(ws, r, 1, name, border=True)
        # Grand Total is a SUM formula across the CURRENT period columns —
        # not a baked-in value — so the operator can edit a cell and
        # the total updates live. (The prior-FY month columns are context
        # and are NOT part of the Grand Total.)
        _cell(ws, r, 2,
              f"=SUM({first_period_col}{r}:{last_period_col}{r})",
              font=_BOLD, fill=_TOTAL_FILL, fmt=INR, border=True)
        # One live SUMIFS per month — the previous months then the
        # selected ones — over the revenue register's Client (G) +
        # Period (J) columns. The FY window shares the "Revenue" sheet
        # with the selected period (v0.3.118) and a month belongs to
        # exactly one of them, so the Period label alone picks the right
        # rows and no Scope criterion is needed; a comparison window
        # reads its own "Revenue (Cmp)" sheet (v0.3.120).
        for i, p in enumerate(fy_months):
            _cell(ws, r, fy0 + i,
                  f'=SUMIFS({prior_rev}!$H:$H,{prior_rev}!$G:$G,$A{r},'
                  f'{prior_rev}!$J:$J,"{month_label(p)}")',
                  fmt=INR, border=True)
        for i, p in enumerate(periods):
            _cell(ws, r, p0 + i,
                  f'=SUMIFS({rev}!$H:$H,{rev}!$G:$G,$A{r},'
                  f'{rev}!$J:$J,"{month_label(p)}")',
                  fmt=INR, border=True)
        r += 1

    # TOTAL row
    last_body = r - 1
    if last_body >= body_start:
        _cell(ws, r, 1, "TOTAL", font=_BOLD, fill=_TOTAL_FILL, border=True)
        for col in range(2, p0 + len(periods)):
            L = get_column_letter(col)
            _cell(ws, r, col,
                  f"=SUM({L}{body_start}:{L}{last_body})",
                  font=_BOLD, fill=_TOTAL_FILL, fmt=INR, border=True)
        _subtotal_filter(ws, hrow, body_start, last_body,
                         list(range(2, p0 + len(periods))),
                         last_col=p0 + len(periods) - 1)

    # Freeze Client + Grand Total; the month columns (prior-FY and
    # current alike) scroll.
    ws.freeze_panes = f"C{body_start}"


def _simple_summary(wb, sheet, title, label, rows, _mapname, lbl, key,
                    sumcol_rev, sumcol_exp, prov_col=None):
    """Entity / Service style summary: name, revenue, expense, net.

    ``prov_col``, when given, names the Provisions-sheet column to match
    this dimension on (Entity = B); outstanding provisions then add into
    the Direct Expense. Service passes ``None`` (provisions carry no
    service dimension)."""
    ws = wb.create_sheet(sheet)
    ws.sheet_view.showGridLines = False
    headers = [label, "Revenue", "Direct Expense", "Net"]
    for i, w in enumerate([30, 16, 16, 16]):
        ws.column_dimensions[get_column_letter(1 + i)].width = w
    _cell(ws, 1, 1, title, font=_TITLE)
    hrow = 3
    _header_row(ws, hrow, headers)

    rev, exp, prov = _q("Revenue"), _q("Expenses"), _q("Provisions")
    # These sheets report the SELECTED period, so exclude the FY-prior
    # rows the data sheets also carry since v0.3.118.
    rev_s = _scope_crit(rev, SCOPE_COL_REV)
    exp_s = _scope_crit(exp, SCOPE_COL_EXP)
    r = hrow + 1
    first = r
    for item in rows:
        _cell(ws, r, 1, item["name"], border=True)
        # Revenue Amount = col H; Expenses Amount = col J (v0.3.69 layout —
        # Invoice No + Type of Expense pushed Amount from H to J; col H is
        # now the "Type of Expense" TEXT column, so summing it returns 0).
        _cell(ws, r, 2,
              f"=SUMIFS({rev}!$H:$H,"
              f"{rev}!${sumcol_rev}:${sumcol_rev},$A{r}{rev_s})",
              fmt=INR, border=True)
        exp_f = (f"=SUMIFS({exp}!$J:$J,"
                 f"{exp}!${sumcol_exp}:${sumcol_exp},$A{r}{exp_s})")
        if prov_col:
            exp_f += f"+SUMIFS({prov}!$G:$G,{prov}!${prov_col}:${prov_col},$A{r})"
        _cell(ws, r, 3, exp_f, fmt=INR, border=True)
        _cell(ws, r, 4, f"=B{r}-C{r}", fmt=INR, border=True)
        r += 1
    if r > first:
        _cell(ws, r, 1, "TOTAL", font=_BOLD, fill=_TOTAL_FILL, border=True)
        for col in (2, 3, 4):
            L = get_column_letter(col)
            _cell(ws, r, col, f"=SUM({L}{first}:{L}{r - 1})", font=_BOLD,
                  fill=_TOTAL_FILL, fmt=INR, border=True)
        _subtotal_filter(ws, hrow, first, r - 1, [2, 3, 4], last_col=4)


# --- data sheets -------------------------------------------------------------

# Data sheets carry a filtered-SUBTOTAL row on top (row 1), the header on
# row 2, and data from row 3 — so applying a column filter shows the live
# sum of the visible rows in the top cell. Full-column SUMIFS elsewhere are
# unaffected: their criteria columns are text, so the subtotal/header rows
# never match a criterion and are excluded from the sum.
_DATA_FIRST_ROW = 3


def _write_data_sheet(wb, name, headers, widths, rows):
    ws = wb.create_sheet(name)
    ncols = len(headers)
    # Detect amount/numeric columns: a column whose data cells are all
    # numbers or formula strings (no plain text). Those get a SUBTOTAL on top.
    has_num = [False] * ncols
    has_text = [False] * ncols
    for row in rows:
        for ci in range(ncols):
            val = row[ci] if ci < len(row) else None
            if val in (None, ""):
                continue
            if isinstance(val, (int, float)) or (
                    isinstance(val, str) and val.startswith("=")):
                has_num[ci] = True
            else:
                has_text[ci] = True
    numeric = [has_num[i] and not has_text[i] for i in range(ncols)]

    _header_row(ws, 2, headers)
    for i, w in enumerate(widths):
        ws.column_dimensions[get_column_letter(1 + i)].width = w
    for ri, row in enumerate(rows, start=_DATA_FIRST_ROW):
        for ci, val in enumerate(row, start=1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.font = _NORMAL
            if isinstance(val, (int, float)):
                c.number_format = INR if abs(val) >= 100 else HOURS
            elif isinstance(val, str) and val.startswith("="):
                # Formula cells (e.g. the Salary sheet's live Overhead
                # amounts) render with the money format.
                c.number_format = INR

    data_last = len(rows) + _DATA_FIRST_ROW - 1
    if rows:
        label_done = False
        for ci in range(ncols):
            L = get_column_letter(ci + 1)
            if numeric[ci]:
                # Open-ended to the sheet bottom (v0.3.117) — a row the
                # operator types in BELOW the generated data still counts.
                # Bounding at data_last silently excluded manual
                # additions. (Rows appended past the AutoFilter range
                # can't be hidden by it, so SUBTOTAL always includes
                # them — coherent with "everything I added counts".)
                cell = ws.cell(
                    row=1, column=ci + 1,
                    value=f"=SUBTOTAL(109,{L}{_DATA_FIRST_ROW}:{L}1048576)")
                cell.font = _BOLD
                cell.fill = _TOTAL_FILL
                cell.number_format = INR
            elif not label_done:
                _cell(ws, 1, ci + 1, "Subtotal (filtered) →", font=_BOLD)
                label_done = True
    ws.freeze_panes = f"A{_DATA_FIRST_ROW}"
    ws.auto_filter.ref = f"A2:{get_column_letter(ncols)}{max(2, data_last)}"
    return ws


def _subtotal_filter(ws, hrow, first_data, last_data, amount_cols, last_col,
                     label_col=1):
    """Add a filtered-SUBTOTAL row just above a summary table's header and
    turn the header + body into an AutoFilter, so filtering the rows shows
    the live sum of the visible rows in that top cell. ``amount_cols`` are
    the 1-based columns to subtotal (percentage columns are skipped). The
    table's own bottom TOTAL row stays a plain SUM (all rows)."""
    if last_data < first_data:
        return
    srow = hrow - 1
    _cell(ws, srow, label_col, "Subtotal (filtered) →", font=_BOLD)
    for col in amount_cols:
        L = get_column_letter(col)
        _cell(ws, srow, col,
              f"=SUBTOTAL(109,{L}{first_data}:{L}{last_data})",
              font=_BOLD, fill=_TOTAL_FILL, fmt=INR)
    ws.auto_filter.ref = (f"{get_column_letter(label_col)}{hrow}:"
                          f"{get_column_letter(last_col)}{last_data}")


def _client_or_party(lbl: dict, f: dict) -> str:
    """Client column value: master canonical name when the party is
    linked; otherwise the raw party name straight off the voucher (the
    vendor/client as Tally spells it). "(unmapped)" only when the
    voucher carried no party at all."""
    name = lbl["cli"].get(f.get("client_id"))
    if name:
        return name
    return (f.get("party_name") or "").strip() or "(unmapped)"


def _sheet_revenue(wb, data: MISData, lbl: dict, suffix: str = "",
                   fy_data: MISData | None = None,
                   scope: str = SCOPE_CURRENT) -> None:
    # Period (J, v0.3.109) lets the Client Billing sheet split the FY
    # months before the selected period into one live SUMIFS column per
    # month (was a single cumulative column).
    #
    # Scope (K, v0.3.118) marks which window a row belongs to. With
    # *fy_data*, that period's rows are written FIRST (they are the
    # earlier months) as "FY Prior", then the selected period's as
    # "Current" — one continuous register instead of the old separate
    # "Revenue (FY)" sheet. See the SCOPE_* constants for why the
    # current-window criterion is ``<>FY Prior``.
    def _row(f, scope):
        svc_name = lbl["svc"].get(f["service_id"], "(unspecified)")
        return [
            _fmt_date(f.get("txn_date")),
            f.get("vch_no") or "",
            lbl["ent"].get(f["entity_id"], "(unspecified)"),
            lbl["cc"].get(f["cost_centre_id"], "Unassigned"),
            lbl["mgr"].get(f["manager_id"], "(unassigned)"),
            svc_name,
            _client_or_party(lbl, f),
            round(f["amount"], 2),
            _service_category(svc_name),
            month_label(f.get("period")) if f.get("period") else "",
            scope,
        ]
    rows = [_row(f, SCOPE_FY) for f in (fy_data.revenue_facts
                                        if fy_data is not None else [])]
    rows += [_row(f, scope) for f in data.revenue_facts]
    _write_data_sheet(
        wb, "Revenue" + suffix,
        ["Date", "Invoice No", "Entity", "CostCentre", "Manager",
         "Service", "Client", "Amount", "Category", "Period", SCOPE_HEADER],
        [12, 16, 24, 12, 12, 22, 28, 14, 14, 10, 11], rows)


def _sheet_expenses(wb, data: MISData, lbl: dict, suffix: str = "",
                    fy_data: MISData | None = None,
                    scope: str = SCOPE_CURRENT) -> None:
    # Column layout (v0.3.69, + M in v0.3.118):
    #   A Date          B Voucher No   C Invoice No   D Entity
    #   E CostCentre    F Manager      G Service      H Type of Expense
    #   I Client        J Amount       K Description  L Period
    #   M Scope
    #
    # Invoice No comes from the register's "New Ref" bill allocation
    # (the vendor's invoice number on the detailed Tally export).
    # Type of Expense bifurcates Professional Fees (direct cost) from
    # Indirect Expense — the P&Ls SUMIFS against column H.
    # Period (L) lets the Employee Register compute the per-period
    # office indirect pool with a SUMIFS.
    # Scope (M) marks the FY-prior rows merged in from what used to be the
    # "Expenses (FY)" sheet — see the SCOPE_* constants.
    def _row(f, scope):
        svc_name = lbl["svc"].get(f["service_id"], "(unspecified)")
        return [
            _fmt_date(f.get("txn_date")), f.get("vch_no") or "",
            f.get("invoice_no") or "",
            lbl["ent"].get(f["entity_id"], "(unspecified)"),
            lbl["cc"].get(f["cost_centre_id"], "Unassigned"),
            lbl["mgr"].get(f["manager_id"], "(unassigned)"),
            svc_name,
            f.get("expense_type") or expense_type(svc_name),
            _client_or_party(lbl, f),
            round(f["amount"], 2), f.get("description", ""),
            month_label(f.get("period")) if f.get("period") else "",
            scope,
        ]
    rows = [_row(f, SCOPE_FY) for f in (fy_data.expense_facts
                                        if fy_data is not None else [])]
    rows += [_row(f, scope) for f in data.expense_facts]
    _write_data_sheet(wb, "Expenses" + suffix,
                      ["Date", "Voucher No", "Invoice No", "Entity",
                       "CostCentre", "Manager", "Service", "Type of Expense",
                       "Client", "Amount", "Description", "Period",
                       SCOPE_HEADER],
                      [12, 14, 16, 24, 12, 12, 20, 17, 28, 14, 30, 10, 11],
                      rows)


def _sheet_provisions(wb, data: MISData, lbl: dict, suffix: str = "") -> None:
    """Outstanding provision costs carried into the reporting month.

    Always created (even with no rows) so the Cost Centre / Entity /
    Partner-Manager P&L SUMIFS over the Remaining column resolve to 0
    rather than #REF. ``CostCentre`` holds the partner code and ``Entity``
    the entity name, matching the criteria those P&L sheets use.
    ``Period`` (A) is the month the provision was BOOKED.

    A provision is a carried-forward stock, not a flow, so each window's
    edition of this sheet holds ONE snapshot: the balance outstanding at
    the end of that window. The **As At** column (H, v0.3.120) names that
    month, so a reader can see which balance they are looking at. No
    formula filters on it — every window has its own sheet — which keeps
    a row the operator appends counting, per v0.3.117.
    """
    default_asat = (max(data.options.periods)
                    if data.options.periods else None)
    rows = []
    for f in data.provision_facts:
        asat = f.get("asat_period") or f.get("period") or default_asat
        rows.append([
            month_label(f.get("provision_period"))
            if f.get("provision_period") else "",
            f.get("entity_name") or "(unspecified)",
            lbl["cc"].get(f["cost_centre_id"], "Unassigned"),
            f.get("client_name") or "(unmapped)",
            round(f.get("original", 0.0), 2),
            round(f.get("adjusted", 0.0), 2),
            round(f["amount"], 2),
            month_label(asat) if asat else "",
        ])
    _write_data_sheet(
        wb, "Provisions" + suffix,
        ["Period", "Entity", "CostCentre", "Client",
         "Original", "Adjusted", "Remaining", "As At"],
        [12, 24, 14, 28, 16, 16, 16, 12], rows)


def _fmt_date(raw):
    """Render a txn_date (stored as ``YYYY-MM-DD`` ISO string) as
    ``DD-Mon-YYYY`` for readability — falls back to whatever's in the
    field if we can't parse it."""
    if not raw:
        return ""
    try:
        return _dt.date.fromisoformat(str(raw)[:10]).strftime("%d-%b-%Y")
    except (ValueError, TypeError):
        return str(raw)


def _sheet_reimbursements(wb, data: MISData, lbl: dict,
                            suffix: str = "") -> None:
    """Per-row reimbursement detail. One row per uploaded entry.

    Columns: Period | Date | Client CC | Employee | Employee CC | Manager |
             Client | Client Reimbursable | Amount | Pool Source

    ``Employee CC`` (E) is the BOOKED cost centre — the employee's own
    home cost centre from the master (v0.3.109), falling back to the
    client's partner, then Office, when the employee isn't resolved yet.
    ``Client CC`` (C) is informational: the partner serving the client
    this was spent for. The two differ when an employee of one partner
    incurs a reimbursement for another partner's client.

    Cost Centre P&L / Partner-Manager P&L SUMIFS this sheet's Amount
    column (I) via **Employee CC (E)** — the cost follows the employee.

    ``Pool Source`` (J, v0.3.109) is "Yes" on rows of office-home staff
    within the selected locations: their outlays join the office-overhead
    pool, and the Employee Register's Office Staff Reimbursement column
    SUMIFS this column, keeping the overhead cascade formula-driven.
    """
    def _client_label(f):
        cid = f.get("client_id")
        if cid is not None:
            return lbl["cli"].get(cid, "(unmapped)")
        raw = f.get("client_raw")
        if raw:
            return f"{raw}  ← unmapped, link in Review tab"
        return "(no client)"
    rows = [[
        month_label(f["period"]), _fmt_date(f.get("txn_date")),
        lbl["cc"].get(f.get("client_cost_centre_id"), "—"),
        f["employee_name"],
        lbl["cc"].get(f["cost_centre_id"], "Unassigned"),
        lbl["mgr"].get(f.get("employee_manager_id"), "(unassigned)"),
        _client_label(f),
        "Yes" if f["client_reimbursable"] else "No",
        round(f["amount"], 2),
        "Yes" if f.get("is_pool_source") else "",
    ] for f in data.reimbursement_facts]
    _write_data_sheet(
        wb, "Reimbursements" + suffix,
        ["Period", "Date", "Client CC", "Employee", "Employee CC", "Manager",
         "Client", "Client Reimbursable", "Amount", "Pool Source"],
        [10, 12, 12, 26, 12, 16, 28, 16, 14, 12], rows)


def _sheet_salary(wb, data: MISData, lbl: dict, suffix: str = "") -> None:
    def _client_label(f):
        # Distinct labels for residual vs unresolved-timesheet vs resolved.
        if f.get("is_overhead_offset"):
            return "(office indirect expenses allocated to employees)"
        if f.get("is_overhead"):
            return "(office overhead per active employee)"
        if f.get("is_residual"):
            return "(residual / unallocated time)"
        cid = f.get("client_id")
        if cid is not None:
            return lbl["cli"].get(cid, "(unmapped)")
        raw = f.get("client_raw")
        if raw:
            return f"{raw}  ← unmapped, link in Review tab"
        return "(no client)"
    # Column layout (v0.3.68 — three CC columns for clarity):
    #   A Period         B Date           C Charged To      D Employee
    #   E Client         F Client CC      G Home CC         H Hours
    #   I Amount         J Type
    #
    # "Charged To" (was "CostCentre" pre-v0.3.68) = where the cost
    # actually lands. SUMIFS in Cost Centre P&L / Partner-Manager P&L
    # uses this column (still at C) so existing formulas keep working;
    # the rename is header-only.
    # "Client CC" — the CC the worked-on client belongs to; blank for
    # residual / non-billable / overhead rows. Helps the operator see
    # WHY a billable slice was charged to a partner other than the
    # employee's home.
    # "Home CC" — the employee's home CC from the master. Always
    # populated. Lets the operator read the salary bifurcation at a
    # glance: an employee whose Home CC differs from Charged To has
    # worked on a client belonging to another partner.
    # Type — "Salary" or "Overhead"; SUMIFS uses this (now at J) to
    # split Salary Cost from Allocated Overhead.
    #
    # Overhead amounts are LIVE formulas chaining to the Employee
    # Register sheet (per-employee share = office indirect pool ÷
    # active count), so tweaking an office expense row recomputes the
    # whole overhead cascade. The comparison-period sheet keeps plain
    # values — its Employee Register isn't part of the workbook.
    def _cc_label(cc_id):
        if cc_id is None:
            return ""
        return lbl["cc"].get(cc_id, "")
    er = _q("Employee Register")
    live = (suffix == "")

    # Row span of each offset row's own heads block. The heads of one
    # period are written contiguously with their offset row right after
    # (order guaranteed by _build_labour_facts), so the offset can back
    # out exactly what those rows charge: -SUM(heads). With the Employee
    # Register's Consider dropdowns (v0.3.106) a what-if flip changes the
    # per-head share (col J) on every head row AND the offset together,
    # keeping the firm total balanced — the old −(share × recipients)
    # form drifted when recipients moved but the written head rows
    # didn't.
    offset_span: dict[int, tuple[int, int]] = {}
    _block_start = None
    for _i, _f in enumerate(data.labour_facts):
        if _f.get("is_overhead") and not _f.get("is_overhead_offset"):
            if _block_start is None:
                _block_start = _i
        else:
            if _f.get("is_overhead_offset") and _block_start is not None:
                offset_span[_i + _DATA_FIRST_ROW] = (
                    _block_start + _DATA_FIRST_ROW,
                    _i - 1 + _DATA_FIRST_ROW)
            _block_start = None

    def _amount(f, r):
        # Overhead amounts chain LIVE to the Employee Register: per-employee
        # share = pool ÷ recipients (col K); the offset backs out the sum
        # of its own heads block. Office indirect stays a SUMIFS so editing
        # an office expense still recomputes the cascade.
        if live and f.get("is_overhead_offset"):
            span = offset_span.get(r)
            if span:
                return f"=-SUM($I${span[0]}:$I${span[1]})"
            return (f"=-SUMIFS({er}!$K:$K,{er}!$A:$A,$A{r})"
                    f"*SUMIFS({er}!$I:$I,{er}!$A:$A,$A{r})")
        if live and f.get("is_overhead"):
            return f"=SUMIFS({er}!$K:$K,{er}!$A:$A,$A{r})"
        return round(f["amount"], 2)

    def _billable(f):
        # Blank for overhead rows; Yes/No for salary so the Partner-Manager
        # P&L can split billable vs non-billable labour.
        if f.get("is_overhead"):
            return ""
        return "Yes" if f.get("billable") else "No"

    # "Pool Source" (M, v0.3.98) — "Yes" on the salary rows of office-home
    # staff (within the selected locations) whose pay feeds the overhead
    # pool. The Employee Register's Office Staff Salary column SUMIFS this
    # column, so the whole overhead cascade stays formula-driven.
    rows = [[month_label(f["period"]), _fmt_date(f.get("txn_date")),
             _cc_label(f["cost_centre_id"]) or "Unassigned",
             f["employee_name"], _client_label(f),
             _cc_label(f.get("client_cost_centre_id")),
             _cc_label(f.get("home_cost_centre_id")),
             round(f["hours"], 2), _amount(f, i + _DATA_FIRST_ROW),
             "Overhead" if f.get("is_overhead") else "Salary",
             _billable(f),
             lbl["mgr"].get(f.get("manager_id"), "(unassigned)"),
             "Yes" if (f.get("is_pool_source")
                       and not f.get("is_overhead")) else "",
             f.get("location") or "—"]
            for i, f in enumerate(data.labour_facts)]
    _write_data_sheet(wb, "Salary" + suffix,
                      ["Period", "Date", "Charged To", "Employee", "Client",
                       "Client CC", "Home CC", "Hours", "Amount", "Type",
                       "Billable", "Manager", "Pool Source", "Location"],
                      [10, 12, 12, 26, 28, 12, 12, 12, 14, 12, 10, 16, 12,
                       14], rows)


# --- Employee Register --------------------------------------------------------

def _sheet_employee_register(wb: Workbook, data: MISData, lbl: dict) -> None:
    """Active headcount per period + joiners/exits, and the office-overhead
    computation the Salary sheet's Overhead rows chain to.

    Three blocks, all formula-driven off the roster table:

    1. **Summary** — one row per period: active employees, new joiners,
       exits (COUNTIFS over the roster), the office indirect pool
       (SUMIFS over the Expenses sheet), the office staff salary and
       reimbursement pool legs (SUMIFS over the Salary / Reimbursements
       sheets' Pool Source columns) and the per-employee overhead
       (pool ÷ recipients). The Salary sheet's Overhead rows SUMIFS into
       column K here, so the whole overhead cascade is live.
    2. **Roster** — one row per (period, employee): home cost centre,
       Active/Exited status, and the movement vs the previous month.
       "Active" = filed timesheet hours in the period. Exited = was in
       the previous month's timesheet, absent this month.
    3. **Headcount by cost centre** — COUNTIFS per (CC, period).
    """
    reg = data.employee_register
    if not reg:
        return
    ws = wb.create_sheet("Employee Register")
    ws.sheet_view.showGridLines = False
    for col, w in (("A", 12), ("B", 26), ("C", 18), ("D", 13), ("E", 13),
                   ("F", 20), ("G", 22)):
        ws.column_dimensions[col].width = w

    office_code = next(
        (c["code"] for c in lbl["cc_active"] if c["cc_type"] == "office"),
        "Office")
    exp = _q("Expenses")
    lab = _q("Salary")
    rmb = _q("Reimbursements")
    # Salary-only row window on the Salary sheet. Salary facts are written
    # first and Overhead facts appended after them (order guaranteed by
    # _build_labour_facts), so bounding the Office-Staff-Salary SUMIFS to
    # this window keeps it clear of the Overhead rows — whose live
    # formulas chain back to THIS sheet's column J (a full-column
    # reference would make Excel flag a circular dependency).
    n_sal = sum(1 for f in data.labour_facts if not f.get("is_overhead"))
    sal_last_row = _DATA_FIRST_ROW + n_sal - 1

    _cell(ws, 1, 1, "Employee Register", font=_TITLE)
    _cell(ws, 2, 1,
          "Active = filed timesheet hours in the period (21st → 20th "
          "cycle). The roster and every headcount (Active / New Joiners / "
          "Exits and the by-cost-centre block) cover EVERYONE, whatever "
          "locations were selected. The Consider column (dropdown) gates "
          "ONLY the office-overhead computation: Overhead Recipients "
          "counts \"Consider\" rows, so flipping a dropdown recomputes "
          "the recipients and per-head figures live (the pool's salary "
          "and reimbursement legs follow the Pool Source columns on the "
          "Salary / Reimbursements sheets). Overhead pool = Office "
          "indirect expenses + office-home staff salaries + office-home "
          "staff reimbursements; it is spread over the considered "
          "partner-team employees AND partners (Recipients) and charged "
          "to their cost centres. The Salary sheet's Overhead rows read "
          "the per-employee figure (column K).",
          font=_SUB)

    # ---- Pre-compute the roster rows so the summary COUNTIFS know
    # their range before being written. Partners (from the cost-centre
    # master, not the employee master) get status "Partner": they count
    # as overhead recipients but not as Active employees (v0.3.98).
    # Roster layout (v0.3.99 — Location; v0.3.106 — Consider):
    #   A Period  B Employee  C Home CC  D Location  E Status  F Movement
    #   G Consider
    def _consider(member) -> str:
        return ("Consider" if member.get("considered", True)
                else "Don't consider")

    roster_rows: list[list] = []
    for r in reg:
        for emp in r["active"]:
            roster_rows.append([
                month_label(r["period"]), emp["name"], emp["cc_code"],
                emp.get("location", "—"), "Active",
                "New Joiner" if emp["is_new"] else "", _consider(emp)])
        for emp in r["exits"]:
            roster_rows.append([
                month_label(r["period"]), emp["name"], emp["cc_code"],
                emp.get("location", "—"), "Exited", "Exit", _consider(emp)])
        for p in r.get("partners", []):
            roster_rows.append([
                month_label(r["period"]), p["name"], p["cc_code"],
                p.get("location", "—"), "Partner", "", _consider(p)])

    sum_hrow = 4
    sum_first = sum_hrow + 1
    sum_last = sum_first + len(reg) - 1
    roster_hrow = sum_last + 2
    roster_first = roster_hrow + 1
    roster_last = max(roster_first, roster_first + len(roster_rows) - 1)
    rA = f"$A${roster_first}:$A${roster_last}"        # Period
    rC = f"$C${roster_first}:$C${roster_last}"        # Home CC
    rStat = f"$E${roster_first}:$E${roster_last}"     # Status
    rMov = f"$F${roster_first}:$F${roster_last}"      # Movement
    rCons = f"$G${roster_first}:$G${roster_last}"     # Consider

    # ---- 1. Summary -----------------------------------------------------
    for col, w in (("F", 18), ("G", 20), ("H", 22), ("I", 16), ("J", 16),
                   ("K", 20)):
        ws.column_dimensions[col].width = w
    _header_row(ws, sum_hrow, [
        "Period", "Prev Month", "Active Employees", "New Joiners", "Exits",
        "Office Indirect (₹)", "Office Staff Salary (₹)",
        "Office Staff Reimbursement (₹)",
        "Overhead Recipients", "Overhead Pool (₹)", "Overhead / Employee (₹)"])
    notes = []
    for i, r in enumerate(reg):
        row = sum_first + i
        _cell(ws, row, 1, month_label(r["period"]), border=True)
        _cell(ws, row, 2, month_label(r["prev_period"])
              + ("" if r["has_prev_data"] else "  (no data)"), border=True)
        # Headcounts cover the WHOLE roster (v0.3.107) — the Consider flag
        # gates only the overhead computation below.
        _cell(ws, row, 3,
              f'=COUNTIFS({rA},$A{row},{rStat},"Active")',
              font=_BOLD, border=True, align=_CENTER)
        _cell(ws, row, 4,
              f'=COUNTIFS({rA},$A{row},{rMov},"New Joiner")',
              border=True, align=_CENTER)
        _cell(ws, row, 5,
              f'=COUNTIFS({rA},$A{row},{rMov},"Exit")',
              border=True, align=_CENTER)
        # Office indirect — live SUMIFS so editing an office expense
        # recomputes the cascade. The Period criterion (col L) already
        # pins this to the selected month, and the FY-prior rows merged
        # into the Expenses sheet in v0.3.118 are all OTHER months, so
        # no Scope criterion is needed here.
        _cell(ws, row, 6,
              f'=SUMIFS({exp}!$J:$J,{exp}!$E:$E,"{office_code}",'
              f'{exp}!$H:$H,"{EXPENSE_TYPE_INDIRECT}",{exp}!$L:$L,$A{row})',
              fmt=INR, border=True)
        # Office-home staff salary — live SUMIFS over the Salary sheet's
        # Pool Source rows (v0.3.98; was a baked value). The range is
        # BOUNDED to the salary-only rows: the Overhead rows that follow
        # them chain back to this sheet's column K, so a full-column
        # reference would be a circular dependency in Excel. This is the
        # ONE bounded window a manually-appended row can't reach (a
        # salary row typed below the Overhead block with Pool Source =
        # Yes won't join the pool) — unavoidable without the circularity;
        # every other figure in the workbook picks appended rows up.
        if sal_last_row >= _DATA_FIRST_ROW:
            sal_rng = lambda col: (f"{lab}!${col}${_DATA_FIRST_ROW}:"
                                   f"${col}${sal_last_row}")
            _cell(ws, row, 7,
                  f'=SUMIFS({sal_rng("I")},{sal_rng("A")},$A{row},'
                  f'{sal_rng("M")},"Yes")',
                  fmt=INR, border=True)
        else:
            _cell(ws, row, 7, 0, fmt=INR, border=True)
        # Office staff reimbursement (v0.3.109) — live SUMIFS over the
        # Reimbursements sheet's Pool Source rows (office-home staff
        # within the selected locations). Static values there, so the
        # full-column reference can't go circular.
        _cell(ws, row, 8,
              f'=SUMIFS({rmb}!$I:$I,{rmb}!$A:$A,$A{row},'
              f'{rmb}!$J:$J,"Yes")',
              fmt=INR, border=True)
        # Overhead recipients — live COUNTIFS over the roster (v0.3.98;
        # was a baked value): partner-team Active employees (Active minus
        # office-home Active) plus the Partner rows — "Consider" rows
        # only (v0.3.106).
        _cell(ws, row, 9,
              f'=COUNTIFS({rA},$A{row},{rStat},"Active",{rCons},"Consider")'
              f'-COUNTIFS({rA},$A{row},{rStat},"Active",{rC},"{office_code}",'
              f'{rCons},"Consider")'
              f'+COUNTIFS({rA},$A{row},{rStat},"Partner",{rCons},"Consider")',
              border=True, align=_CENTER)
        # Pool = office indirect + office staff salary + office staff
        # reimbursement.
        _cell(ws, row, 10, f"=F{row}+G{row}+H{row}", fmt=INR, border=True)
        # Per-employee overhead = pool ÷ recipients.
        _cell(ws, row, 11,
              f"=IF(I{row}=0,0,IF(J{row}<=0,0,ROUND(J{row}/I{row},2)))",
              font=_BOLD, fmt=INR, border=True)
        if not r["has_prev_data"]:
            notes.append(
                f"{month_label(r['period'])}: no timesheet found for "
                f"{month_label(r['prev_period'])} — "
                f"joiners/exits not computed for this period.")

    # ---- 2. Roster -------------------------------------------------------
    _header_row(ws, roster_hrow, [
        "Period", "Employee", "Home CC", "Location", "Status", "Movement",
        "Consider"])
    for i, row_vals in enumerate(roster_rows):
        row = roster_first + i
        for ci, val in enumerate(row_vals, start=1):
            _cell(ws, row, ci, val, border=True,
                  align=_CENTER if ci in (1, 3, 4, 5, 6, 7) else None)
    # Consider column is a two-value dropdown (v0.3.106) so the operator
    # can flip a row in/out of the computation — every COUNTIFS above
    # carries the "Consider" criterion, so the counts, recipients and
    # per-head overhead recompute live from the flip.
    if roster_rows:
        dv = DataValidation(
            type="list", formula1='"Consider,Don\'t consider"',
            allow_blank=False, showDropDown=False)
        dv.error = 'Pick "Consider" or "Don\'t consider".'
        dv.errorTitle = "Invalid value"
        ws.add_data_validation(dv)
        dv.add(f"G{roster_first}:G{roster_last}")

    # ---- 3. Headcount by cost centre -------------------------------------
    # Placed BESIDE the summary (top-right) rather than at the bottom of
    # the sheet, so the operator reads both summary tables at a glance
    # without scrolling past the full roster. Starts at column M — one gap
    # after the summary's overhead columns (which now run through K).
    cc_codes: list[str] = []
    for rr in roster_rows:
        if rr[2] not in cc_codes:
            cc_codes.append(rr[2])
    cc_codes.sort()
    if cc_codes:
        hc_c0 = 13                               # column M (one gap after K)
        ccL = get_column_letter(hc_c0)           # Cost Centre column
        perL = get_column_letter(hc_c0 + 1)      # Period column
        for off, w in enumerate((20, 12, 10, 13, 10)):
            ws.column_dimensions[get_column_letter(hc_c0 + off)].width = w
        _cell(ws, sum_hrow - 1, hc_c0, "Headcount by cost centre", font=_BOLD)
        _header_row(ws, sum_hrow, [
            "Cost Centre", "Period", "Active", "New Joiners", "Exits"],
            start_col=hc_c0)
        row = sum_hrow + 1
        for code in cc_codes:
            for r in reg:
                _cell(ws, row, hc_c0, code, border=True)
                _cell(ws, row, hc_c0 + 1, month_label(r["period"]),
                      border=True, align=_CENTER)
                # Whole-roster headcounts (v0.3.107) — no Consider filter.
                _cell(ws, row, hc_c0 + 2,
                      f'=COUNTIFS({rA},${perL}{row},{rC},${ccL}{row},'
                      f'{rStat},"Active")',
                      border=True, align=_CENTER)
                _cell(ws, row, hc_c0 + 3,
                      f'=COUNTIFS({rA},${perL}{row},{rC},${ccL}{row},'
                      f'{rMov},"New Joiner")',
                      border=True, align=_CENTER)
                _cell(ws, row, hc_c0 + 4,
                      f'=COUNTIFS({rA},${perL}{row},{rC},${ccL}{row},'
                      f'{rMov},"Exit")',
                      border=True, align=_CENTER)
                row += 1

    for i, note in enumerate(notes):
        _cell(ws, roster_last + 2 + i, 1, "• " + note, font=_SUB)

    ws.freeze_panes = f"A{sum_first}"


# --- Client Register ---------------------------------------------------------

def _sheet_client_register(wb: Workbook, data: MISData, lbl: dict,
                           prior: "Window | None" = None) -> None:
    """Clients billed per period, flagging only NEW clients (first ever
    billed) — the client analogue of the Employee Register. Clients bill
    irregularly, so a month's absence is NOT treated as "lost"; the system
    remembers every past client and only additions are flagged. Summary
    COUNTIFS over a per-(period, client) roster, plus a 'Clients by cost
    centre' breakdown beside it.

    The register covers the PREVIOUS months as well (v0.3.120, operator
    ask): a Jul-26 MIS lists Apr-26, May-26 and Jun-26 above it, so the
    board reads the year's client movement in one place. Those are the
    same FY-so-far months (*prior*) the Partner-Manager P&L's cumulative
    column covers. A Scope column tells the two apart.
    """
    prior_months = (set(prior.periods)
                    if prior is not None and prior.active else set())
    prior_reg = ([r for r in prior.data.client_register
                  if r["period"] in prior_months] if prior_months else [])
    reg = sorted(prior_reg + list(data.client_register),
                 key=lambda r: r["period"])
    if not reg:
        return
    scope_of = {r["period"]: ("Previous" if r["period"] in prior_months
                              else "Reported") for r in reg}
    ws = wb.create_sheet("Client Register")
    ws.sheet_view.showGridLines = False
    for col, w in (("A", 12), ("B", 32), ("C", 16), ("D", 10)):
        ws.column_dimensions[col].width = w

    _cell(ws, 1, 1, "Client Register", font=_TITLE)
    _cell(ws, 2, 1,
          "Active = billed on a sales voucher in the period. New = billed "
          "for the FIRST time ever (across all months on record). Clients "
          "bill irregularly, so a month's absence isn't shown as lost — the "
          "system keeps remembering every past client. Grouped by the "
          "client's cost centre."
          + ("  The 'Previous' months — the earlier months of this "
             "financial year — are listed alongside the reported one so "
             "the year's client movement reads in one place."
             if prior_months else ""), font=_SUB)

    # Pre-compute the roster so the summary COUNTIFS know their range.
    roster_rows: list[list] = []
    for r in reg:
        for c in r["active"]:
            roster_rows.append([
                month_label(r["period"]), c["name"], c["cc_code"],
                "New" if c["is_new"] else ""])

    sum_hrow = 4
    sum_first = sum_hrow + 1
    sum_last = sum_first + len(reg) - 1
    roster_hrow = sum_last + 2
    roster_first = roster_hrow + 1
    roster_last = max(roster_first, roster_first + len(roster_rows) - 1)
    rA = f"$A${roster_first}:$A${roster_last}"
    rC = f"$C${roster_first}:$C${roster_last}"
    rD = f"$D${roster_first}:$D${roster_last}"

    # ---- 1. Summary -----------------------------------------------------
    # Scope (B) is a label, not a criterion — every count stays keyed on
    # the Period alone, exactly as before.
    _header_row(ws, sum_hrow,
                ["Period", "Scope", "Active Clients", "New Clients"])
    for i, r in enumerate(reg):
        row = sum_first + i
        _cell(ws, row, 1, month_label(r["period"]), border=True)
        _cell(ws, row, 2, scope_of[r["period"]], border=True, align=_CENTER)
        # Active = every roster row for the period; New = the flagged ones.
        _cell(ws, row, 3, f'=COUNTIFS({rA},$A{row})',
              font=_BOLD, border=True, align=_CENTER)
        _cell(ws, row, 4, f'=COUNTIFS({rA},$A{row},{rD},"New")',
              border=True, align=_CENTER)

    # ---- 2. Roster -------------------------------------------------------
    _header_row(ws, roster_hrow, ["Period", "Client", "Cost Centre", "New?"])
    for i, row_vals in enumerate(roster_rows):
        row = roster_first + i
        for ci, val in enumerate(row_vals, start=1):
            _cell(ws, row, ci, val, border=True,
                  align=_CENTER if ci in (1, 4) else None)

    # ---- 3. Clients by cost centre (beside the summary) ------------------
    cc_codes: list[str] = []
    for rr in roster_rows:
        if rr[2] not in cc_codes:
            cc_codes.append(rr[2])
    cc_codes.sort()
    if cc_codes:
        c0 = 6                                   # column F (one gap after D)
        ccL = get_column_letter(c0)              # Cost Centre column
        perL = get_column_letter(c0 + 1)         # Period column
        for off, w in enumerate((16, 12, 10, 10)):
            ws.column_dimensions[get_column_letter(c0 + off)].width = w
        _cell(ws, sum_hrow - 1, c0, "Clients by cost centre", font=_BOLD)
        _header_row(ws, sum_hrow, [
            "Cost Centre", "Period", "Active", "New"], start_col=c0)
        row = sum_hrow + 1
        for code in cc_codes:
            for r in reg:
                _cell(ws, row, c0, code, border=True)
                _cell(ws, row, c0 + 1, month_label(r["period"]),
                      border=True, align=_CENTER)
                _cell(ws, row, c0 + 2,
                      f'=COUNTIFS({rA},${perL}{row},{rC},${ccL}{row})',
                      border=True, align=_CENTER)
                _cell(ws, row, c0 + 3,
                      f'=COUNTIFS({rA},${perL}{row},{rC},${ccL}{row},'
                      f'{rD},"New")', border=True, align=_CENTER)
                row += 1

    ws.freeze_panes = f"A{sum_first}"


# --- Comparatives ------------------------------------------------------------

def _sheet_comparatives(wb: Workbook, data: MISData, lbl: dict,
                        windows: "Windows", rows_pl: dict) -> None:
    """Year to date vs the same period last year — revenue & profit per
    cost centre.

    The cost-centre companion to "Partner-Manager P&L (Cmp)", on the same
    two windows (v0.3.121; it used to compare the selected period against
    an operator-ticked one). Both sides are built by the SAME formula
    against their own data sheets — " (YTD)" and " (LY)" — so the two
    columns are like for like, and the whole sheet is live.
    """
    ws = wb.create_sheet("Comparatives")
    ws.sheet_view.showGridLines = False
    cur_lbl = windows.ytd.label
    cmp_lbl = windows.last_year.label
    headers = ["Code", "Cost Centre",
               f"Revenue — {cur_lbl}", f"Revenue — {cmp_lbl}", "Revenue Δ",
               f"Profit — {cur_lbl}", f"Profit — {cmp_lbl}", "Profit Δ",
               "Profit Δ %"]
    for i, w in enumerate([10, 24, 18, 18, 16, 18, 18, 16, 11]):
        ws.column_dimensions[get_column_letter(1 + i)].width = w

    _cell(ws, 1, 1, "Year on Year — cost centre comparison", font=_TITLE)
    _cell(ws, 2, 1,
          f"Financial year to date: {cur_lbl}    vs    the same period last "
          f"year: {cmp_lbl}.  Reporting period: "
          f"{periods_label(windows.selected)}.",
          font=_SUB)
    hrow = 4
    _header_row(ws, hrow, headers)

    def revenue(sfx: str, row: int) -> str:
        # Total income: every Revenue row for the cost centre (sales plus
        # the reimbursement / OPE recoveries).
        rev = _data_sheet_q("Revenue", sfx)
        return f"=SUMIFS({rev}!$H:$H,{rev}!$D:$D,$A{row})"

    def profit(sfx: str, row: int, revenue_cell: str) -> str:
        # Net profit = total income − expenses (professional + indirect)
        # − labour (salary AND the Overhead-type rows, which carry the
        # per-employee office allocation and the Office offset) −
        # reimbursement outlays − outstanding provisions. Same definition
        # as the Cost Centre P&L's Net Profit column, expressed against
        # one window's sheets so both sides here are like for like.
        exp = _data_sheet_q("Expenses", sfx)
        lab = _q("Salary" + sfx)
        reimb, prov = _q("Reimbursements" + sfx), _q("Provisions" + sfx)
        return (f"={revenue_cell}"
                f"-SUMIFS({exp}!$J:$J,{exp}!$E:$E,$A{row})"
                f"-SUMIFS({lab}!$I:$I,{lab}!$C:$C,$A{row})"
                f"-SUMIFS({reimb}!$I:$I,{reimb}!$E:$E,$A{row})"
                f"-SUMIFS({prov}!$G:$G,{prov}!$C:$C,$A{row})")

    pl = _q("Cost Centre P&L")
    first = hrow + 1
    r = first
    for code, plrow in rows_pl["row_of"].items():
        _cell(ws, r, 1, code, border=True)
        _cell(ws, r, 2, f"={pl}!B{plrow}", border=True)
        _cell(ws, r, 3, revenue(YTD, r), fmt=INR, border=True)
        _cell(ws, r, 4, revenue(LY, r), fmt=INR, border=True)
        _cell(ws, r, 5, f"=C{r}-D{r}", fmt=INR, border=True)
        _cell(ws, r, 6, profit(YTD, r, f"C{r}"), fmt=INR, border=True)
        _cell(ws, r, 7, profit(LY, r, f"D{r}"), fmt=INR, border=True)
        _cell(ws, r, 8, f"=F{r}-G{r}", fmt=INR, border=True)
        _cell(ws, r, 9, f"=IF(G{r}=0,\"\",(F{r}-G{r})/ABS(G{r}))",
              fmt=PCT, border=True, align=_CENTER)
        r += 1

    last = r - 1
    _cell(ws, r, 2, "TOTAL", font=_BOLD, fill=_TOTAL_FILL, border=True)
    for col in (3, 4, 5, 6, 7, 8):
        L = get_column_letter(col)
        _cell(ws, r, col, f"=SUM({L}{first}:{L}{last})", font=_BOLD,
              fill=_TOTAL_FILL, fmt=INR, border=True)
    _cell(ws, r, 9, f"=IF(G{r}=0,\"\",(F{r}-G{r})/ABS(G{r}))", font=_BOLD,
          fill=_TOTAL_FILL, fmt=PCT, border=True, align=_CENTER)
    _subtotal_filter(ws, hrow, first, last, [3, 4, 5, 6, 7, 8], last_col=9)
