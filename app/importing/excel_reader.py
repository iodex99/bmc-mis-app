"""Read uploaded Excel files into plain Python row grids.

Both modern `.xlsx` (openpyxl) and legacy `.xls` (xlrd) are supported; Tally
exports its sales register as `.xls` so the latter is required.
"""

from __future__ import annotations

import datetime as _dt
from pathlib import Path
from typing import Any

import openpyxl
import xlrd


def _is_xls(path: str | Path) -> bool:
    return str(path).lower().endswith(".xls")


def sheet_names(path: str | Path) -> list[str]:
    """List worksheet names in a workbook."""
    if _is_xls(path):
        wb = xlrd.open_workbook(path, on_demand=True)
        try:
            return list(wb.sheet_names())
        finally:
            wb.release_resources()
    wb = openpyxl.load_workbook(path, read_only=True)
    try:
        return list(wb.sheetnames)
    finally:
        wb.close()


def read_grid(path: str | Path, sheet: str | None = None) -> list[list[Any]]:
    """Return a worksheet as a list of equal-length rows of cell values."""
    if _is_xls(path):
        rows = _read_xls(path, sheet)
    else:
        rows = _read_xlsx(path, sheet)
    width = max((len(r) for r in rows), default=0)
    for r in rows:
        r.extend([None] * (width - len(r)))
    return rows


def _read_xlsx(path, sheet) -> list[list[Any]]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb[sheet] if sheet else wb.worksheets[0]
        return [list(r) for r in ws.iter_rows(values_only=True)]
    finally:
        wb.close()


def _read_xls(path, sheet) -> list[list[Any]]:
    wb = xlrd.open_workbook(path)
    try:
        ws = wb.sheet_by_name(sheet) if sheet else wb.sheet_by_index(0)
        datemode = wb.datemode
        out: list[list[Any]] = []
        for r in range(ws.nrows):
            row: list[Any] = []
            for c in range(ws.ncols):
                cell = ws.cell(r, c)
                if cell.ctype == xlrd.XL_CELL_DATE:
                    row.append(xlrd.xldate_as_datetime(cell.value, datemode))
                elif cell.ctype == xlrd.XL_CELL_EMPTY:
                    row.append(None)
                elif cell.ctype == xlrd.XL_CELL_BOOLEAN:
                    row.append(bool(cell.value))
                else:
                    row.append(cell.value)
            out.append(row)
        return out
    finally:
        wb.release_resources()


def guess_header_row(grid: list[list[Any]], must_contain: tuple[str, ...] = ()) -> int:
    """Best-guess the index of the header row.

    Picks the row with the most non-empty text cells; if *must_contain* keywords
    are given, prefers a row containing one of them.
    """
    best_idx, best_score = 0, -1
    for i, row in enumerate(grid[:40]):
        texts = [str(c).strip().lower() for c in row if c not in (None, "")]
        if not texts:
            continue
        score = sum(1 for c in row if isinstance(c, str) and c.strip())
        if must_contain and any(k in t for t in texts for k in must_contain):
            score += 100
        if score > best_score:
            best_idx, best_score = i, score
    return best_idx


def excel_serial_to_date(value: float) -> _dt.date | None:
    """Convert an Excel date serial (e.g. ``46024``) to a date."""
    try:
        # Excel epoch quirk: 1900 is treated as a leap year, hence Dec 30, 1899.
        return (_dt.datetime(1899, 12, 30) + _dt.timedelta(days=float(value))).date()
    except (ValueError, OverflowError):
        return None
